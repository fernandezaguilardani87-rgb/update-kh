#!/usr/bin/env python3
"""
Property Manager — KH Inmobiliaria Costa del Sol
Gestor automatizado del listado maestro de propiedades

Uso: python3 property_manager.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from functools import lru_cache
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from datetime import datetime
from pathlib import Path

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    import fitz          # pymupdf — convierte páginas PDF a imagen
    from PIL import Image
    import pytesseract
    import io as _io
    import platform as _platform
    import os as _os

    # Ruta de Tesseract en Windows (auto-detección extendida)
    if _platform.system() == 'Windows':
        _local_app    = _os.environ.get('LOCALAPPDATA', '')
        _user_profile = _os.environ.get('USERPROFILE', '')
        _tesseract_paths = [
            r'C:\Program Files\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
            _os.path.join(_local_app,    'Programs', 'Tesseract-OCR', 'tesseract.exe'),
            _os.path.join(_local_app,    'Tesseract-OCR', 'tesseract.exe'),
            _os.path.join(_user_profile, 'AppData', 'Local', 'Programs',
                          'Tesseract-OCR', 'tesseract.exe'),
            _os.path.join(_user_profile, 'AppData', 'Local',
                          'Tesseract-OCR', 'tesseract.exe'),
        ]
        for _tp in _tesseract_paths:
            if _tp and Path(_tp).exists():
                pytesseract.pytesseract.tesseract_cmd = _tp
                break

    # Validar que Tesseract funciona de verdad (no solo que se importó pytesseract)
    pytesseract.get_tesseract_version()
    OCR_SUPPORT = True
    OCR_ERROR   = ''
except Exception as _ocr_ex:
    OCR_SUPPORT = False
    OCR_ERROR   = str(_ocr_ex)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

APP_TITLE = "Property Manager — KH Inmobiliaria"
VERSION = "1.0"

MASTER_COLUMNS = [
    'Promotor', 'Development', 'Municipality', 'Reference',
    'Bedrooms', 'Price', 'Floor', 'Information', 'Map', 'Delivery',
    'Observations'
]

INTERNAL_COLS = ['_status', '_change']  # Not exported to Excel header row

COL_WIDTHS_APP = {
    'Promotor': 160, 'Development': 155, 'Municipality': 100, 'Reference': 125,
    'Bedrooms': 72, 'Price': 105, 'Floor': 110, 'Information': 120,
    'Map': 50, 'Delivery': 90, 'Observations': 230
}

COL_WIDTHS_XL = {
    'Promotor': 22, 'Development': 24, 'Municipality': 16, 'Reference': 20,
    'Bedrooms': 10, 'Price': 14, 'Floor': 18, 'Information': 20,
    'Map': 8, 'Delivery': 14, 'Observations': 40
}

STATUS_CHANGED  = 'changed'   # → fila amarillo suave
STATUS_SOLD     = 'sold'      # → fila gris claro + mover al final
STATUS_NO_PRICE = 'no_price'  # → precio en rojo suave (fila rosa en app, celda en Excel)
STATUS_NONE     = ''

# ── Modern UI color palette ───────────────────────────────────────────────────
# Backgrounds
BG_APP      = '#F1F5F9'    # Slate-100 — ventana principal
BG_SIDEBAR  = '#0F172A'    # Slate-900 — sidebar oscuro
BG_TOOLBAR  = '#1E293B'    # Slate-800 — barra de herramientas
BG_CARD     = '#FFFFFF'    # Blanco — tabla / cards
BG_STRIPE   = '#F8FAFC'    # Slate-50 — filas alternas
# Accent & semantic
C_ACCENT    = '#6366F1'    # Indigo-500 — acento principal
C_SUCCESS   = '#10B981'    # Emerald-500 — ok / actualizar
C_INFO      = '#3B82F6'    # Blue-500 — exportar / ver
C_WARNING   = '#F59E0B'    # Amber-500 — informe
C_DANGER    = '#EF4444'    # Red-500 — eliminar
C_ADD       = '#8B5CF6'    # Violet-500 — añadir
C_EDIT      = '#0EA5E9'    # Sky-500 — editar
# Text
T_MAIN      = '#0F172A'    # texto principal
T_MUTED     = '#64748B'    # texto secundario
T_SIDE      = '#CBD5E1'    # texto sidebar
T_SIDE_MUTED= '#475569'    # texto sidebar suave
# Status row colors (tabla)
COLOR_YELLOW = '#FEF9C3'   # amber-100 — precio modificado
COLOR_GRAY   = '#E2E8F0'   # slate-200 — vendida
COLOR_PINK   = '#FEE2E2'   # red-100 — sin precio
# Row text for sold
T_SOLD      = '#94A3B8'    # slate-400

# Fills para Excel (sin cambios — afectan el archivo guardado)
FILL_YELLOW   = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
FILL_GRAY     = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
FILL_RED_CELL = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
FILL_HEADER   = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')

# Font stack — Segoe UI (Windows), SF Pro (Mac), system sans-serif fallback
FONT_UI   = 'Segoe UI'
FONT_MONO = 'Cascadia Code'

# Orden de prioridad para mostrar (los sold van al fondo)
STATUS_ORDER = {STATUS_NONE: 0, STATUS_CHANGED: 1, STATUS_NO_PRICE: 2, STATUS_SOLD: 3}

# Rango de precio válido para propiedades en Costa del Sol.
# Cualquier número fuera de este rango se descarta como error de OCR / referencia numérica.
PRICE_MIN =    50_000   # 50 k€
PRICE_MAX = 10_000_000  # 10 M€ (villas de lujo extremo)

SCRIPT_DIR   = Path(__file__).parent
LISTADOS_DIR = Path(r"C:\Users\User\Docs\Update\Update KH\Listados de Precios")

# ─────────────────────────────────────────────────────────────────────────────
# FLOOR CLASSIFICATION
# ─────────────────────────────────────────────────────────────────────────────

def classify_floor(floor_str: str) -> str:
    """Classify floor label into: ground | middle | penthouse | villa | other"""
    f = str(floor_str).lower().strip()
    if any(w in f for w in ['bajo', 'ground', 'planta baja', 'level 0', 'nivel 0',
                             'baja', 'pb', 'planta 0', 'sótano', 'sotano']):
        return 'ground'
    # GF / BJ / B0 = Ground Floor
    if re.match(r'^(gf|bj|b0|bajo\s*\w*)$', f):
        return 'ground'
    # Floor 0 o "0" numérico puro = planta baja
    if re.match(r'^0+$', f):
        return 'ground'
    if any(w in f for w in ['pent', 'ático', 'atico', 'solarium', 'duplex pent', 'atic']):
        return 'penthouse'
    # AT / Atico alone
    if re.match(r'^(at|at\s*\w?|ático|atico)$', f):
        return 'penthouse'
    if any(w in f for w in ['villa', 'semi-detach', 'townhouse', 'attached', 'plot']):
        return 'villa'
    if any(w in f for w in ['1st', '2nd', '3rd', '4th', '1er', '2do', '3er',
                             'primero', 'segundo', 'tercero', 'cuarto',
                             '1ª', '2ª', '3ª', 'middle', 'intermedia',
                             '1st floor', '2d floor', 'level 1', 'level 2']):
        return 'middle'
    if 'floor' in f or 'planta' in f:
        return 'middle'
    # Número puro (1, 2, 3…) = planta intermedia
    if re.match(r'^\d+$', f):
        return 'middle'
    return 'other'


def _floor_from_ref(ref: str) -> str:
    """
    Deriva la etiqueta de planta desde el código de referencia de la unidad.
    Fallback para PDFs sin columna PLANTA explícita (p.ej. Prime Invest).

    Patrones soportados:
      AE-1.02.A   → "2"      (bloque.planta.unidad con puntos)
      AE-2.00.D   → "0"      (planta baja numérica)
      AE-1.S01.A  → "Sótano"
      360-10A     → "10"     (prefijo-2dígitos+letra)
      AS-11B      → "11"
      AN-24A      → "2"      (primer dígito como planta)
      AN-2        → "2"
    """
    ref = str(ref).strip()
    if not ref:
        return ''

    # Patrón 1: PREFIX-block.FLOOR.unit  (ej. AE-1.02.A, AE-2.00.D, AE-1.S01.A)
    m = re.search(r'\d+\.([^.]+)\.[^.]+$', ref)
    if m:
        seg = m.group(1)
        prefix = re.sub(r'\d', '', seg).upper()
        digits = re.sub(r'\D', '', seg).lstrip('0') or '0'
        if prefix in ('S', 'SS', 'SOT'):
            return 'Sótano'
        if prefix in ('A', 'AT', 'PH'):
            return 'Ático'
        return digits

    # Patrón 2: PREFIX-{2+digits}{letter(s)}  (ej. 360-10A, AS-11B, AS-40A)
    m = re.search(r'-(\d{2,})[A-Za-z]+$', ref)
    if m:
        return m.group(1).lstrip('0') or '0'

    # Patrón 3: PREFIX-{digit}{digit}{letter+}  (ej. AN-24A → piso 2)
    m = re.search(r'-(\d)(\d[A-Za-z]+)$', ref)
    if m:
        return m.group(1)

    # Patrón 4: PREFIX-{digits}  (ej. AN-2, AN-11 sin letra final)
    m = re.search(r'-(\d+)$', ref)
    if m:
        return m.group(1)

    # Patrón 5: Último recurso — primer dígito(s) justo tras cualquier guión
    # (ej. 360-4-1A → "4", AL4-2113 → "2113")
    m = re.search(r'-(\d+)', ref)
    if m:
        return m.group(1)

    return ''


# ─────────────────────────────────────────────────────────────────────────────
# PRICE LIST DETECTION
# ─────────────────────────────────────────────────────────────────────────────

KEYWORDS_PRECIO = [
    'price list', 'pricelist', 'price-list', 'price_list',
    'lista de precio', 'lista precios', 'listado de precio',
    'lista_de_precio', 'lista_precios', 'tarifa', 'precios',
]

def es_listado_precios(nombre: str) -> bool:
    """True si el nombre del archivo parece un listado de precios."""
    n = nombre.lower().replace('_', ' ').replace('-', ' ')
    if re.search(r'\blp\b', n):
        return True
    return any(kw.replace('_', ' ').replace('-', ' ') in n for kw in KEYWORDS_PRECIO)


def parse_price(value) -> float | None:
    """
    Convierte un precio a float.
    Maneja todos los formatos europeos y anglosajones:
      1.234.567 €   → 1234567   (puntos = separadores de miles)
      1,234,567 €   → 1234567   (comas = separadores de miles)
      1.234,56 €    → 1234.56   (europeo: punto=miles, coma=decimal)
      1,234.56 €    → 1234.56   (anglosajón: coma=miles, punto=decimal)
      399.000 €     → 399000    (punto único con 3 dec. = miles)
      695,000 €     → 695000    (coma única con 3 dec. = miles)
      399,5 €       → 399.5     (coma decimal)
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip().upper()
    if s in ['SOLD', 'RESERVED', 'VENDIDA', 'RESERVADA', 'N/A', '']:
        return None
    s = re.sub(r'[€$£\s]', '', s)
    if not s:
        return None

    ndots   = s.count('.')
    ncommas = s.count(',')

    if ndots > 1 and ncommas == 0:
        s = s.replace('.', '')                      # 1.234.567 → 1234567
    elif ncommas > 1 and ndots == 0:
        s = s.replace(',', '')                      # 1,234,567 → 1234567
    elif ndots == 1 and ncommas == 1:
        if s.index('.') < s.index(','):
            s = s.replace('.', '').replace(',', '.') # 1.234,56 → 1234.56
        else:
            s = s.replace(',', '')                   # 1,234.56 → 1234.56
    elif ndots == 1 and ncommas == 0:
        after = s.split('.')[-1]
        if len(after) == 3 and after.isdigit():
            s = s.replace('.', '')                   # 399.000 → 399000
        # else: 399.5 → decimal, dejar tal cual
    elif ncommas == 1 and ndots == 0:
        after = s.split(',')[-1]
        if len(after) == 3 and after.isdigit():
            s = s.replace(',', '')                   # 695,000 → 695000
        else:
            s = s.replace(',', '.')                  # 399,5 → 399.5

    try:
        result = float(s)
        return result if result > 0 else None
    except ValueError:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXTRACTION — motor multi-estrategia con detección de columnas
# ─────────────────────────────────────────────────────────────────────────────

# Palabras clave por tipo de columna (en varios idiomas)
_PRICE_HDR  = {'price', 'precio', 'pvp', 'importe', 'valor', 'coste', 'cost',
               'value', 'sale price', 'venta', 'total'}
_STATUS_HDR = {'status', 'estado', 'disponibilidad', 'disponib',
               'availability', 'avail', 'situacion', 'situación'}
_BED_HDR    = {'bed', 'dorm', 'hab', 'rooms', 'bedroom', 'dormitorio',
               'habitaciones', 'dormitorios'}
               # NOTA: 'tipo'/'typology' eliminados — son subcadenas de 'tipologia'
               # y causaban que la columna TIPOLOGIA se asignara a dormitorios
               # en lugar de a planta, dejando DORMITORIOS sin mapear.
_FLOOR_HDR  = {'floor', 'planta', 'nivel', 'level', 'plant', 'piso',
               'planta/piso', 'floor/level',
               'tipologia', 'tipolog', 'typology', 'tipo'}
_REF_HDR    = {'ref', 'unit', 'unidad', 'apartment', 'apto', 'piso',
               'vivienda', 'id', 'código', 'code', 'villa', 'número', 'num'}


def _normalize_for_match(text: str) -> str:
    """Normaliza texto para comparación: minúsculas, sin especiales, sin duplicados."""
    t = str(text).lower()
    t = re.sub(r'[^a-z0-9\s]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()


def _fuzzy_word_in_text(word: str, text_tokens: list[str]) -> bool:
    """
    Comprueba si `word` aparece en la lista de tokens del texto, con tolerancia
    a un error tipográfico (distancia de edición ≤ 1) para palabras largas (≥ 6 chars).
    """
    if word in text_tokens or any(word in tok for tok in text_tokens):
        return True
    if len(word) < 6:
        return False
    # Distancia de edición ≤ 1 (inserción / eliminación / sustitución)
    for tok in text_tokens:
        if abs(len(tok) - len(word)) > 1:
            continue
        if len(tok) == len(word):          # Sustitución
            if sum(a != b for a, b in zip(tok, word)) <= 1:
                return True
        else:                              # Inserción / eliminación
            longer, shorter = (tok, word) if len(tok) > len(word) else (word, tok)
            for i in range(len(longer)):
                if longer[:i] + longer[i+1:] == shorter:
                    return True
    return False


# Artículos y preposiciones cortos que no aportan a la identificación de un desarrollo.
_STOP_WORDS_DEV = frozenset({
    'the', 'las', 'los', 'del', 'des', 'and', 'con', 'sin',
    'sur', 'sea', 'san', 'new', 'old', 'von', 'for',
})


@lru_cache(maxsize=2048)
def _dev_score_in_text(dev_name: str, text_norm: str) -> float:
    """
    Puntuación de coincidencia entre un nombre de desarrollo y un texto normalizado.

    Palabras significativas: len ≥ 3 y no en _STOP_WORDS_DEV.
    ── Incluir palabras de 3 caracteres es clave para capturar nombres cortos como
       "Aby" o "Air", que de lo contrario quedarían excluidos y la única palabra
       que puntuaría sería un topónimo compartido ("Estepona"), causando falsas asignaciones.

    Umbral: se requiere que coincidan al menos ⌈60%⌉ de las palabras significativas
    (mínimo 1). Para nombres de 2 palabras esto exige que AMBAS coincidan:
    ── "Aby Estepona" necesita "aby" Y "estepona" en el PDF → "Zenith Estepona" no basta.
    ── "Zenith" (1 palabra) solo necesita "zenith" → sí está en su PDF.

    Fórmula: matched_chars² / total_chars  (favorece cobertura completa del nombre).
    """
    words = [w for w in _normalize_for_match(dev_name).split()
             if len(w) >= 3 and w not in _STOP_WORDS_DEV]
    if not words:
        return 0.0
    tokens = text_norm.split()
    matched_count = sum(1 for w in words if _fuzzy_word_in_text(w, tokens))
    # Requiere ≥ 60% de palabras significativas, mínimo 1
    # ceil(n * 0.6) sin import math: -(-n*3 // 5)
    min_required = max(1, -(-len(words) * 3 // 5))
    if matched_count < min_required:
        return 0.0
    matched_chars = sum(len(w) for w in words if _fuzzy_word_in_text(w, tokens))
    total_chars   = sum(len(w) for w in words)
    # Puntuación cuadrática normalizada
    return float(matched_chars * matched_chars) / float(total_chars)


def _detect_column_map(header_row: list) -> dict:
    """
    Detecta qué columna corresponde a qué dato (precio, estado, dormitorios, etc.)
    a partir de la fila de cabecera de la tabla.
    Devuelve un dict {tipo: índice_columna}.
    """
    col_map = {}
    for i, cell in enumerate(header_row):
        h = str(cell or '').lower().strip()
        if not h:
            continue
        if any(k in h for k in _PRICE_HDR)  and 'price'     not in col_map:
            col_map['price']     = i
        if any(k in h for k in _STATUS_HDR) and 'status'    not in col_map:
            col_map['status']    = i
        # Floor antes que beds: 'tipologia' coincide con 'tipo' en _BED_HDR,
        # pero debe ir a floor. Al evaluar floor primero queda asignado y
        # la rama de beds lo ignora (ya ocupado).
        if any(k in h for k in _FLOOR_HDR)  and 'floor'     not in col_map:
            col_map['floor']     = i
        if any(k in h for k in _BED_HDR)    and 'bedrooms'  not in col_map:
            col_map['bedrooms']  = i
        if any(k in h for k in _REF_HDR)    and 'reference' not in col_map:
            col_map['reference'] = i
    return col_map


def _status_from_text(text: str) -> str:
    t = text.lower()
    if any(w in t for w in ['sold', 'vendid', 'vend ', 'vendida', 'no disponible']):
        return 'SOLD'
    if any(w in t for w in ['reserv', 'reserved', 'option', 'opcion']):
        return 'RESERVED'
    return 'AVAILABLE'


def _parse_row_with_map(row: list, col_map: dict) -> dict | None:
    """Extrae una unidad de una fila usando el mapa de columnas detectado."""
    cells = [str(c).strip() if c else '' for c in row]
    if not any(cells):
        return None

    unit = {}

    # Precio (columna detectada primero, luego búsqueda en toda la fila)
    if 'price' in col_map:
        price = parse_price(cells[col_map['price']])
        if price and PRICE_MIN <= price <= PRICE_MAX:
            unit['price'] = price
    if 'price' not in unit:
        for cell in cells:
            price = parse_price(cell)
            if price and PRICE_MIN <= price <= PRICE_MAX:
                unit['price'] = price
                break
    if 'price' not in unit:
        return None

    # Estado
    status_src = cells[col_map['status']] if 'status' in col_map else ' '.join(cells)
    unit['status'] = _status_from_text(status_src)

    # Dormitorios
    if 'bedrooms' in col_map:
        m = re.search(r'(\d)', cells[col_map['bedrooms']])
        if m:
            unit['bedrooms'] = int(m.group(1))
    if 'bedrooms' not in unit:
        combined = ' '.join(cells)
        m = re.search(r'(\d)\s*(?:dorm|bed|hab|D\b)', combined, re.IGNORECASE)
        if m:
            unit['bedrooms'] = int(m.group(1))

    # Planta
    if 'floor' in col_map:
        unit['floor'] = cells[col_map['floor']]
    if not unit.get('floor'):
        combined = ' '.join(cells).lower()
        for label in ['penthouse', 'ático', 'atico', 'duplex pent', 'ground floor',
                      'planta baja', '1st floor', '2nd floor', '3rd floor',
                      'primero', 'segundo', 'bajo', 'villa', 'townhouse']:
            if label in combined:
                unit['floor'] = label
                break

    # Referencia
    if 'reference' in col_map:
        unit['reference'] = cells[col_map['reference']]
    if not unit.get('reference'):
        combined = ' '.join(cells)
        m = re.search(r'[A-Z]{2,}-[A-Z]{0,3}-?\d+', combined)
        if m:
            unit['reference'] = m.group(0)

    # Fallback: derivar planta desde la referencia si aún no se conoce
    if not unit.get('floor') and unit.get('reference'):
        derived = _floor_from_ref(unit['reference'])
        if derived:
            unit['floor'] = derived

    unit['raw'] = cells
    return unit


def _parse_row_for_unit(row: list) -> dict | None:
    """Parseo heurístico sin mapa de columnas: busca precio > 50k y extrae lo demás."""
    if not row or not any(row):
        return None
    cells = [str(c).strip() if c else '' for c in row]
    combined = ' '.join(cells)

    if not re.search(r'\d{4,}', combined):
        return None

    unit = {}
    for cell in cells:
        price = parse_price(cell)
        if price and PRICE_MIN <= price <= PRICE_MAX:
            unit['price'] = price
            break
    if 'price' not in unit:
        return None

    combined_lower = combined.lower()
    unit['status'] = _status_from_text(combined_lower)

    m = re.search(r'(\d)\s*(?:dorm|bed|hab|D\b)', combined, re.IGNORECASE)
    if m:
        unit['bedrooms'] = int(m.group(1))

    for label in ['penthouse', 'ático', 'atico', 'duplex pent', 'ground floor',
                  'planta baja', '1st floor', '2nd floor', '3rd floor',
                  'primero', 'segundo', 'bajo', 'villa', 'townhouse']:
        if label in combined_lower:
            unit['floor'] = label
            break

    m2 = re.search(r'[A-Z]{2,}-[A-Z]{2,}-\d+', combined)
    if m2:
        unit['reference'] = m2.group(0)

    # Fallback: derivar planta desde la referencia si aún no se conoce
    if not unit.get('floor') and unit.get('reference'):
        derived = _floor_from_ref(unit['reference'])
        if derived:
            unit['floor'] = derived

    unit['raw'] = cells
    return unit


def _parse_text_for_units(text: str) -> list[dict]:
    """Extracción de emergencia desde texto plano cuando no hay tablas."""
    units = []
    for line in text.split('\n'):
        m = re.search(r'(\d{3}[.,]\d{3}(?:[.,]\d{3})?)', line)
        if not m:
            continue
        price_str = m.group(1).replace('.', '').replace(',', '')
        try:
            price = float(price_str)
        except ValueError:
            continue
        if not (PRICE_MIN <= price <= PRICE_MAX):
            continue

        unit = {'price': price, 'status': _status_from_text(line.lower()), 'raw': [line]}
        bm = re.search(r'(\d)\s*(?:dorm|bed|hab)', line.lower())
        if bm:
            unit['bedrooms'] = int(bm.group(1))
        units.append(unit)
    return units


def _parse_ocr_spatial(img) -> list[dict]:
    """
    Parsea una imagen usando coordenadas XY de cada palabra (image_to_data).
    Asocia precios con SOLD/BED/planta solo dentro de un radio espacial estricto,
    evitando contaminación de unidades adyacentes en el mismo rango de línea.

    Radio: 400px horizontal × 100px vertical (separa portales y plantas distintas).
    Prioridad de planta: referencia de apartamento (1A, 1D, 0A) > palabras clave.
    """
    data = pytesseract.image_to_data(
        img, config='--psm 11 --oem 3',
        output_type=pytesseract.Output.DICT)

    tokens: list[dict] = []
    for i, word in enumerate(data['text']):
        word = word.strip()
        if not word or int(data['conf'][i]) < 15:
            continue
        tokens.append({
            'text': word,
            'cx': data['left'][i] + data['width'][i] // 2,
            'cy': data['top'][i] + data['height'][i] // 2,
        })

    _price_re = re.compile(r'[€£$]?\s*(\d{1,3}(?:[.,\s]\d{3})+)')
    _bed_re   = re.compile(r'(\d)\s*(?:BED|DORM(?:ITORIOS?)?|HAB(?:ITACIONES?)?|BR\b)',
                           re.IGNORECASE)
    _ref_re   = re.compile(r'\b(\d[A-Za-z])\b')   # 1A, 1D, 0A, 2B …

    RADIUS_X = 400   # px — separa los dos portales de un mismo bloque
    RADIUS_Y = 100   # px — separa plantas adyacentes (tipicamente 120-130 px)

    units: list[dict] = []
    seen_prices: set[float] = set()

    for tok in tokens:
        pm = _price_re.search(tok['text'])
        if not pm:
            continue
        raw = pm.group(1).replace(' ', '')
        if raw.count('.') > 1 or (
                raw.count('.') == 1 and raw.count(',') == 0
                and len(raw.split('.')[-1]) == 3):
            raw = raw.replace('.', '').replace(',', '')
        else:
            raw = raw.replace(',', '').replace('.', '')
        try:
            price = float(raw)
        except ValueError:
            continue
        if not (PRICE_MIN <= price <= PRICE_MAX) or price in seen_prices:
            continue

        px, py = tok['cx'], tok['cy']
        nearby = [t for t in tokens if t is not tok
                  and abs(t['cx'] - px) <= RADIUS_X
                  and abs(t['cy'] - py) <= RADIUS_Y]
        nearby_text = ' '.join(t['text'] for t in nearby).lower()

        # ── Estado ─────────────────────────────────────────────────────
        if 'sold' in nearby_text or 'vendid' in nearby_text:
            status = 'SOLD'
        elif 'reserv' in nearby_text or 'option' in nearby_text:
            status = 'RESERVED'
        else:
            status = 'AVAILABLE'

        unit: dict = {
            'price': price, 'status': status, 'ocr': True,
            'raw': [tok['text']] + [t['text'] for t in nearby],
        }

        # ── Dormitorios ────────────────────────────────────────────────
        bm = _bed_re.search(nearby_text)
        if bm:
            unit['bedrooms'] = int(bm.group(1))

        # ── Planta: referencia de apartamento (mayor prioridad) ────────
        floor = ''
        rm = _ref_re.search(nearby_text)
        if rm:
            ref = rm.group(1).upper()
            if re.match(r'^0[A-Z]', ref):
                floor = 'Ground Floor'
            elif re.match(r'^[1-9][A-Z]', ref):
                floor = '1st floor'
        # Palabras clave como fallback (si no hay referencia)
        if not floor:
            nt = nearby_text
            if any(w in nt for w in ['penthouse', 'atico', 'ático', 'solarium', 'duplex pent']):
                floor = 'Penthouse'
            elif any(w in nt for w in ['bajo', 'ground floor', 'planta baja', 'gf']):
                floor = 'Ground Floor'
            elif any(w in nt for w in ['1st floor', '2nd floor', '1st', '2nd', '3rd']):
                floor = '1st floor'
        if floor:
            unit['floor'] = floor

        units.append(unit)
        seen_prices.add(price)

    return units


def _parse_ocr_text(text: str) -> list[dict]:
    """
    Parsea el texto extraído por OCR de una página de imagen.
    Busca precios (€XXX.XXX / €XXX,XXX), dormitorios (X BED/DORM) y estado (SOLD).
    Optimizado para renders aéreos tipo Bromley con cajas de precio superpuestas.
    """
    lines  = [l.strip() for l in text.split('\n') if l.strip()]
    units  = []

    # Formatos de precio: €710,000 / €710.000 / 710.000 / 710,000 / 710 000
    price_re = re.compile(
        r'[€£$]?\s*(\d{1,3}(?:[.,\s]\d{3})+)',
        re.IGNORECASE)
    # Alternativa: número largo sin separador visible (OCR puede omitir comas)
    price_re2 = re.compile(r'\b(\d{6,7})\b')

    bed_re   = re.compile(r'(\d)\s*(?:BED|DORM(?:ITORIOS?)?|HAB(?:ITACIONES?)?|BR\b)', re.IGNORECASE)
    ref_re   = re.compile(r'\b(\d[A-Z](?:-?\d+)?|\w{2,4}-\d{2,4})\b')  # 0A, 1B, 2C, IH-101…

    seen_prices: set[float] = set()
    all_text = text.lower()

    for i, line in enumerate(lines):
        # Intento 1: precio con separadores de miles
        pm = price_re.search(line)
        raw_num = None
        if pm:
            raw_num = pm.group(1).replace(' ', '')
            # Normalizar separadores: si el punto separa miles → eliminar
            if raw_num.count('.') > 1 or (
                    raw_num.count('.') == 1 and raw_num.count(',') == 0
                    and len(raw_num.split('.')[-1]) == 3):
                raw_num = raw_num.replace('.', '').replace(',', '')
            else:
                raw_num = raw_num.replace(',', '').replace('.', '')
        else:
            # Intento 2: número de 6-7 dígitos sin separadores (OCR a veces los omite)
            pm2 = price_re2.search(line)
            if pm2:
                raw_num = pm2.group(1)

        if not raw_num:
            continue

        try:
            price = float(raw_num)
        except ValueError:
            continue

        if not (PRICE_MIN <= price <= PRICE_MAX) or price in seen_prices:
            continue

        # Ventana de contexto amplia (±4 líneas) para detectar SOLD cerca del precio
        ctx_lines = lines[max(0, i - 4): min(len(lines), i + 5)]
        ctx       = ' '.join(ctx_lines)
        ctx_lower = ctx.lower()

        # Estado: buscar SOLD / VENDIDA en la misma línea primero (alta confianza)
        line_lower = line.lower()
        if 'sold' in line_lower or 'vendid' in line_lower:
            status = 'SOLD'
        elif 'sold' in ctx_lower or 'vendid' in ctx_lower:
            status = 'SOLD'
        elif 'reserv' in ctx_lower or 'option' in ctx_lower:
            status = 'RESERVED'
        else:
            status = 'AVAILABLE'

        unit: dict = {'price': price, 'status': status, 'raw': ctx_lines, 'ocr': True}

        bm = bed_re.search(ctx)
        if bm:
            unit['bedrooms'] = int(bm.group(1))

        rm = ref_re.search(ctx)
        ref_str = ''
        if rm:
            ref_str = rm.group(1)
            unit['reference'] = ref_str

        # Inferir planta desde referencia de apartamento Bromley:
        # 0A/0B = Ground, 1A/1B/2A/2B = Intermedia, AT/ATx = Ático
        if ref_str and not unit.get('floor'):
            if re.match(r'^AT', ref_str, re.I):
                unit['floor'] = 'Penthouse'
            elif re.match(r'^0[A-Z]', ref_str):
                unit['floor'] = 'Ground Floor'
            elif re.match(r'^[1-9][A-Z]', ref_str):
                unit['floor'] = '1st floor'

        # Inferir planta desde palabras clave en el contexto
        if not unit.get('floor'):
            ctx_l = ctx_lower
            if any(w in ctx_l for w in ['penthouse', 'atico', 'ático', 'solarium', 'duplex pent']):
                unit['floor'] = 'Penthouse'
            elif any(w in ctx_l for w in ['ground floor', 'planta baja', 'bajo', 'gf']):
                unit['floor'] = 'Ground Floor'
            elif any(w in ctx_l for w in ['1st floor', '2nd floor', '1st', '2nd']):
                unit['floor'] = '1st floor'

        units.append(unit)
        seen_prices.add(price)

    return units


def _extract_ocr(pdf_path: str) -> tuple[list[dict], str]:
    """
    Fallback OCR: convierte cada página a imagen (alta resolución) y
    pasa Tesseract para extraer texto de PDFs basados en imagen.
    Devuelve (units, error_msg). Si error_msg != '' hubo un problema.

    Estrategia multi-PSM:
      - PSM 11 (sparse text): ideal para renders con texto disperso
      - PSM 6  (uniform block): segunda pasada si PSM 11 da poco
    """
    if not OCR_SUPPORT:
        return [], f'OCR no disponible: {OCR_ERROR}'

    units: list[dict] = []

    try:
        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            page  = doc[page_num]
            # Alta resolución (~300 DPI) para texto pequeño en renders
            mat   = fitz.Matrix(3.0, 3.0)
            pix   = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            img   = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)

            # Prioridad 1: parser espacial (usa coordenadas XY — evita contaminación
            # entre unidades adyacentes en el mismo rango horizontal de texto)
            page_units = _parse_ocr_spatial(img)

            # Prioridad 2: parser de texto PSM 11 si el espacial no da resultados
            if not page_units:
                text11 = pytesseract.image_to_string(
                    img, config='--psm 11 --oem 3 -c preserve_interword_spaces=1')
                page_units = _parse_ocr_text(text11)

            # Prioridad 3: PSM 6 como último recurso
            if not page_units:
                text6 = pytesseract.image_to_string(
                    img, config='--psm 6 --oem 3 -c preserve_interword_spaces=1')
                page_units = _parse_ocr_text(text6)

            units.extend(page_units)
        doc.close()
    except Exception as e:
        return units, f'Error OCR: {e}'

    return units, ''


def extract_units_from_pdf(pdf_path: str) -> list[dict]:
    """
    Extrae unidades de precio de un PDF con estrategia multi-nivel:
      1. Detección de cabecera de columna — tablas estructuradas (más preciso)
      2. Parseo heurístico fila a fila
      3. Extracción desde texto plano (fallback texto)
      4. OCR via Tesseract (fallback para PDFs de imagen)
    """
    if not PDF_SUPPORT:
        return []

    units: list[dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_units: list[dict] = []
            tables = page.extract_tables()

            for table in (tables or []):
                if not table or len(table) < 2:
                    continue

                # Estrategia 1: cabecera detectada
                col_map = _detect_column_map(table[0])
                if col_map and 'price' in col_map:
                    for row in table[1:]:
                        u = _parse_row_with_map(row, col_map)
                        if u:
                            page_units.append(u)
                else:
                    # Estrategia 2: heurístico
                    for row in table:
                        u = _parse_row_for_unit(row)
                        if u:
                            page_units.append(u)

            # Estrategia 3: texto plano si las tablas no dieron nada
            if not page_units:
                text = page.extract_text() or ''
                page_units.extend(_parse_text_for_units(text))

            units.extend(page_units)

    # Estrategia 4: OCR si pdfplumber no extrajo nada (PDF de imagen)
    ocr_error = ''
    if not units:
        units, ocr_error = _extract_ocr(pdf_path)

    # Guardar error de OCR en primer elemento si aplica (para mostrarlo en el log)
    if ocr_error and not units:
        return [{'_ocr_error': ocr_error}]

    # Deduplicar
    seen: set = set()
    deduped: list[dict] = []
    for u in units:
        key = (u.get('price'), u.get('status'), u.get('bedrooms'),
               str(u.get('floor', '')), str(u.get('reference', '')))
        if key not in seen:
            seen.add(key)
            deduped.append(u)

    return deduped


# ─────────────────────────────────────────────────────────────────────────────
# MAIN APPLICATION
# ─────────────────────────────────────────────────────────────────────────────

class PropertyManagerApp:

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1480x860")
        self.root.minsize(960, 640)
        self.root.configure(bg=BG_APP)

        self.df = pd.DataFrame(columns=MASTER_COLUMNS + INTERNAL_COLS)
        self.current_file: str | None = None
        self.changes: list[dict] = []
        self._sort_col = None
        self._sort_asc = True
        self._undo_stack: list[pd.DataFrame] = []   # snapshots para Ctrl+Z

        self._setup_styles()
        self._build_ui()
        self._auto_load()

    # ── Styles ──────────────────────────────────────────────────

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        # ── Treeview (tabla principal) ─────────────────────────────────────
        style.configure('Treeview',
                        rowheight=30,
                        font=(FONT_UI, 10),
                        fieldbackground=BG_CARD,
                        background=BG_CARD,
                        foreground=T_MAIN,
                        borderwidth=0,
                        relief='flat')
        style.configure('Treeview.Heading',
                        font=(FONT_UI, 9, 'bold'),
                        background=BG_TOOLBAR,
                        foreground=T_SIDE,
                        relief='flat',
                        borderwidth=0,
                        padding=(8, 7))
        style.map('Treeview.Heading',
                  background=[('active', '#334155')],
                  foreground=[('active', 'white')])
        style.map('Treeview',
                  background=[('selected', C_ACCENT)],
                  foreground=[('selected', 'white')])
        # ── Scrollbars delgadas ────────────────────────────────────────────
        style.configure('Vertical.TScrollbar',
                        background='#E2E8F0', troughcolor=BG_CARD,
                        borderwidth=0, arrowsize=12, relief='flat')
        style.configure('Horizontal.TScrollbar',
                        background='#E2E8F0', troughcolor=BG_CARD,
                        borderwidth=0, arrowsize=12, relief='flat')
        style.configure('TScrollbar', background='#ddd', troughcolor='#f0f2f5')

    # ── UI Build ─────────────────────────────────────────────────

    def _build_ui(self):
        self._build_menu()
        self._build_toolbar()
        self._build_body()
        self._build_statusbar()

    def _build_menu(self):
        mb = tk.Menu(self.root)

        fm = tk.Menu(mb, tearoff=0)
        fm.add_command(label='Cargar Excel maestro…',        command=self.cmd_load_excel,   accelerator='Ctrl+O')
        fm.add_command(label='↩ Deshacer',                  command=self.cmd_undo,         accelerator='Ctrl+Z')
        fm.add_command(label='🔄 Actualizar Listados (auto)', command=self.cmd_auto_update,  accelerator='Ctrl+U')
        fm.add_command(label='Importar PDFs de promotoras…', command=self.cmd_import_pdfs)
        fm.add_separator()
        fm.add_command(label='Exportar Excel actualizado…', command=self.cmd_export_excel, accelerator='Ctrl+S')
        fm.add_command(label='Generar informe de cambios…', command=self.cmd_generate_report)
        fm.add_separator()
        fm.add_command(label='Salir', command=self.root.quit)
        mb.add_cascade(label='Archivo', menu=fm)

        em = tk.Menu(mb, tearoff=0)
        em.add_command(label='Añadir propiedad',        command=self.cmd_add)
        em.add_command(label='Editar fila seleccionada', command=self.cmd_edit)
        em.add_command(label='Eliminar fila seleccionada', command=self.cmd_delete)
        mb.add_cascade(label='Editar', menu=em)

        vm = tk.Menu(mb, tearoff=0)
        vm.add_command(label='Mostrar todas',              command=lambda: self._apply_filters())
        vm.add_command(label='Solo cambios (amarillo)',   command=lambda: self._apply_filters(status=STATUS_CHANGED))
        vm.add_command(label='Solo sin precio (rosa)',    command=lambda: self._apply_filters(status=STATUS_NO_PRICE))
        vm.add_command(label='Solo vendidas/agotadas',    command=lambda: self._apply_filters(status=STATUS_SOLD))
        mb.add_cascade(label='Vista', menu=vm)

        self.root.config(menu=mb)
        self.root.bind('<Control-o>', lambda e: self.cmd_load_excel())
        self.root.bind('<Control-s>', lambda e: self.cmd_export_excel())
        self.root.bind('<Control-u>', lambda e: self.cmd_auto_update())
        self.root.bind('<Control-z>', lambda e: self.cmd_undo())

    def _build_toolbar(self):
        bar = tk.Frame(self.root, bg=BG_TOOLBAR)
        bar.pack(fill=tk.X)
        # Franja de acento en la parte superior
        tk.Frame(bar, bg=C_ACCENT, height=3).pack(fill=tk.X)

        inner = tk.Frame(bar, bg=BG_TOOLBAR)
        inner.pack(fill=tk.X, padx=10, pady=5)

        def _darken(hex_color: str) -> str:
            h = hex_color.lstrip('#')
            r, g, bl = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
            return f'#{max(0,int(r*.80)):02x}{max(0,int(g*.80)):02x}{max(0,int(bl*.80)):02x}'

        def btn(parent, text, cmd, color):
            b = tk.Button(parent, text=text, command=cmd,
                          bg=color, fg='white',
                          font=(FONT_UI, 9, 'bold'),
                          relief='flat', padx=13, pady=6,
                          cursor='hand2',
                          activebackground=_darken(color),
                          activeforeground='white', bd=0)
            b.pack(side=tk.LEFT, padx=2)
            b.bind('<Enter>', lambda e, c=color: b.configure(bg=_darken(c)))
            b.bind('<Leave>', lambda e, c=color: b.configure(bg=c))
            return b

        def sep():
            tk.Frame(inner, bg='#334155', width=1).pack(
                side=tk.LEFT, fill=tk.Y, padx=6, pady=2)

        btn(inner, '⊞  Cargar Excel',      self.cmd_load_excel,      '#475569')
        sep()
        btn(inner, '⟳  Actualizar',        self.cmd_auto_update,     C_SUCCESS)
        btn(inner, '↑  Importar PDFs',     self.cmd_import_pdfs,     '#059669')
        sep()
        btn(inner, '↓  Exportar Excel',    self.cmd_export_excel,    C_INFO)
        btn(inner, '≡  Informe',           self.cmd_generate_report, C_WARNING)
        sep()
        btn(inner, '+  Añadir',            self.cmd_add,             C_ADD)
        btn(inner, '✎  Editar',            self.cmd_edit,            C_EDIT)
        btn(inner, '✕  Eliminar',          self.cmd_delete,          C_DANGER)

        # ── Búsqueda (derecha) ────────────────────────────────────────────
        sf = tk.Frame(inner, bg='#273548', padx=8, pady=5)
        sf.pack(side=tk.RIGHT, padx=4)
        tk.Label(sf, text='⌕', bg='#273548', fg='#94A3B8',
                 font=(FONT_UI, 12)).pack(side=tk.LEFT, padx=(0,5))
        self._search_var = tk.StringVar()
        self._search_var.trace_add('write', lambda *_: self._on_search())
        se = tk.Entry(sf, textvariable=self._search_var, width=22,
                      font=(FONT_UI, 10), relief='flat',
                      bg='#273548', fg='white',
                      insertbackground='white', highlightthickness=0)
        se.pack(side=tk.LEFT)
        _ph = 'Buscar…'
        def _fi(e):
            if se.get() == _ph: se.delete(0, tk.END); se.configure(fg='white')
        def _fo(e):
            if not se.get(): se.insert(0, _ph); se.configure(fg='#475569')
        se.insert(0, _ph); se.configure(fg='#475569')
        se.bind('<FocusIn>', _fi); se.bind('<FocusOut>', _fo)

    def _build_body(self):
        body = tk.Frame(self.root, bg=BG_APP)
        body.pack(fill=tk.BOTH, expand=True)

        # ══════════════════════════════════════════════════════════════════
        # SIDEBAR OSCURO
        # ══════════════════════════════════════════════════════════════════
        sidebar = tk.Frame(body, bg=BG_SIDEBAR, width=220)
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)

        # Logo / branding
        brand = tk.Frame(sidebar, bg='#0B1120', pady=14)
        brand.pack(fill=tk.X)
        tk.Label(brand, text='KH', font=(FONT_UI, 20, 'bold'),
                 bg='#0B1120', fg=C_ACCENT).pack(side=tk.LEFT, padx=(16, 4))
        tk.Label(brand, text='Property\nManager', font=(FONT_UI, 8),
                 bg='#0B1120', fg=T_SIDE, justify=tk.LEFT).pack(side=tk.LEFT)

        tk.Frame(sidebar, bg='#1E3A5F', height=1).pack(fill=tk.X)

        # Etiqueta sección
        tk.Label(sidebar, text='PROMOTORAS',
                 font=(FONT_UI, 8, 'bold'), bg=BG_SIDEBAR, fg='#475569',
                 anchor='w').pack(fill=tk.X, padx=14, pady=(14, 4))

        # Listbox con fondo oscuro
        lb_wrap = tk.Frame(sidebar, bg=BG_SIDEBAR)
        lb_wrap.pack(fill=tk.BOTH, expand=True, padx=8)

        self._promotor_lb = tk.Listbox(lb_wrap,
                                        font=(FONT_UI, 10),
                                        relief='flat', bd=0,
                                        selectbackground=C_ACCENT,
                                        selectforeground='white',
                                        bg=BG_SIDEBAR, fg=T_SIDE,
                                        activestyle='none',
                                        highlightthickness=0,
                                        exportselection=False)
        lb_scroll = ttk.Scrollbar(lb_wrap, orient=tk.VERTICAL,
                                   command=self._promotor_lb.yview)
        self._promotor_lb.configure(yscrollcommand=lb_scroll.set)
        lb_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self._promotor_lb.pack(fill=tk.BOTH, expand=True)
        self._promotor_lb.bind('<<ListboxSelect>>', self._on_promotor_select)
        self._promotor_lb.bind('<MouseWheel>',
            lambda e: self._promotor_lb.yview_scroll(-1 * (e.delta // 120), 'units'))
        self._promotor_lb.bind('<Button-4>',
            lambda e: self._promotor_lb.yview_scroll(-1, 'units'))   # Linux scroll up
        self._promotor_lb.bind('<Button-5>',
            lambda e: self._promotor_lb.yview_scroll(1, 'units'))    # Linux scroll down

        # Botón "Ver todas"
        tb = tk.Button(sidebar, text='Ver todas',
                       command=lambda: self._apply_filters(),
                       font=(FONT_UI, 9), bg='#1E293B', fg=T_SIDE,
                       relief='flat', cursor='hand2', pady=6,
                       activebackground=C_ACCENT, activeforeground='white',
                       bd=0, highlightthickness=0)
        tb.pack(fill=tk.X, padx=8, pady=(6, 0))

        # Stats card
        tk.Frame(sidebar, bg='#1E293B', height=1).pack(fill=tk.X, pady=(10, 0))
        self._stats_lbl = tk.Label(sidebar, text='',
                                    font=(FONT_UI, 8),
                                    bg='#0B1120', fg='#64748B',
                                    justify=tk.LEFT, anchor='nw',
                                    padx=14, pady=10,
                                    wraplength=195)
        self._stats_lbl.pack(fill=tk.X)

        # ══════════════════════════════════════════════════════════════════
        # ÁREA PRINCIPAL (tabla)
        # ══════════════════════════════════════════════════════════════════
        right = tk.Frame(body, bg=BG_APP)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 0))

        # ── Barra de leyenda ──────────────────────────────────────────────
        legend = tk.Frame(right, bg=BG_CARD, pady=7)
        legend.pack(fill=tk.X)
        tk.Label(legend, text='Estado:', font=(FONT_UI, 9, 'bold'),
                 bg=BG_CARD, fg=T_MUTED).pack(side=tk.LEFT, padx=(14, 10))
        for dot_color, label, status in [
            (COLOR_YELLOW, 'Precio modificado',   STATUS_CHANGED),
            (COLOR_PINK,   'Sin precio',           STATUS_NO_PRICE),
            (COLOR_GRAY,   'Vendida',              STATUS_SOLD),
            (BG_CARD,      'Sin cambios',          STATUS_NONE),
        ]:
            chip = tk.Frame(legend, bg=BG_CARD, cursor='hand2')
            chip.pack(side=tk.LEFT, padx=6)
            dot = tk.Frame(chip, bg=dot_color, width=11, height=11,
                           relief='solid', bd=1)
            dot.pack(side=tk.LEFT, padx=(0, 4))
            lbl = tk.Label(chip, text=label, font=(FONT_UI, 9),
                           bg=BG_CARD, fg=T_MUTED, cursor='hand2')
            lbl.pack(side=tk.LEFT)
            for w in (chip, dot, lbl):
                w.bind('<Button-1>', lambda e, s=status: self._apply_filters(status=s))

        tk.Frame(right, bg='#E2E8F0', height=1).pack(fill=tk.X)

        # ── Tabla ────────────────────────────────────────────────────────
        tframe = tk.Frame(right, bg=BG_CARD)
        tframe.pack(fill=tk.BOTH, expand=True)

        self._tree = ttk.Treeview(tframe, columns=MASTER_COLUMNS,
                                   show='headings', selectmode='browse')
        self._tree.tag_configure('changed',  background=COLOR_YELLOW)
        self._tree.tag_configure('sold',     background=COLOR_GRAY,   foreground=T_SOLD)
        self._tree.tag_configure('no_price', background=COLOR_PINK)
        self._tree.tag_configure('normal',   background=BG_CARD)
        self._tree.tag_configure('stripe',   background=BG_STRIPE)

        for col in MASTER_COLUMNS:
            self._tree.heading(col, text=f'  {col}',
                               command=lambda c=col: self._toggle_sort(c))
            self._tree.column(col, width=COL_WIDTHS_APP.get(col, 100),
                              minwidth=40, stretch=True)

        vsb = ttk.Scrollbar(tframe, orient=tk.VERTICAL,   command=self._tree.yview)
        hsb = ttk.Scrollbar(tframe, orient=tk.HORIZONTAL, command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT,  fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self._tree.pack(fill=tk.BOTH, expand=True)

        self._tree.bind('<Double-1>', lambda _: self.cmd_edit())
        self._tree.bind('<Delete>',   lambda _: self.cmd_delete())

    def _build_statusbar(self):
        bar = tk.Frame(self.root, bg=BG_TOOLBAR, pady=0)
        bar.pack(fill=tk.X, side=tk.BOTTOM)
        tk.Frame(bar, bg='#334155', height=1).pack(fill=tk.X)
        inner = tk.Frame(bar, bg=BG_TOOLBAR)
        inner.pack(fill=tk.X, padx=10, pady=4)
        self._status_var = tk.StringVar(value='Listo — carga tu Excel maestro para comenzar.')
        self._count_var  = tk.StringVar(value='0 propiedades')
        # Dot indicator
        self._status_dot = tk.Label(inner, text='●', bg=BG_TOOLBAR, fg='#10B981',
                                     font=(FONT_UI, 9))
        self._status_dot.pack(side=tk.LEFT, padx=(0, 5))
        tk.Label(inner, textvariable=self._status_var, bg=BG_TOOLBAR, fg=T_SIDE,
                 font=(FONT_UI, 9)).pack(side=tk.LEFT)
        tk.Label(inner, textvariable=self._count_var, bg=BG_TOOLBAR, fg='#475569',
                 font=(FONT_UI, 9)).pack(side=tk.RIGHT)

    # ── Undo ─────────────────────────────────────────────────────

    def _snapshot(self):
        """Guarda una copia del DataFrame antes de una operación destructiva."""
        self._undo_stack.append(self.df.copy())
        if len(self._undo_stack) > 20:          # máximo 20 niveles
            self._undo_stack.pop(0)

    def cmd_undo(self):
        if not self._undo_stack:
            self._status_var.set('Nada que deshacer.')
            return
        self.df = self._undo_stack.pop()
        self._rebuild_search_cache()
        self._refresh_table()
        self._refresh_promotor_list()
        self._update_stats()
        self._status_var.set('Deshacer — estado anterior restaurado.')

    # ── Auto-load ────────────────────────────────────────────────

    def _auto_load(self):
        candidates = [
            SCRIPT_DIR / 'KH.xlsx',
            SCRIPT_DIR / 'kh.xlsx',
        ]
        for p in candidates:
            if p.exists():
                self._load_excel(str(p))
                return

    # ── Data Loading ─────────────────────────────────────────────

    def cmd_load_excel(self):
        path = filedialog.askopenfilename(
            title='Cargar Excel maestro',
            initialdir=str(SCRIPT_DIR),
            filetypes=[('Excel', '*.xlsx *.xls'), ('Todos', '*.*')])
        if path:
            self._load_excel(path)

    def _load_excel(self, path: str):
        try:
            raw = pd.read_excel(path, sheet_name=0, header=None, dtype=str)

            # Find header row containing 'Promotor'
            header_row = None
            for i, row in raw.iterrows():
                if 'Promotor' in row.values:
                    header_row = i
                    break

            if header_row is None:
                messagebox.showerror('Error',
                    'No se encontró la fila de encabezados "Promotor" en el Excel.')
                return

            df = pd.read_excel(path, sheet_name=0, header=header_row, dtype=str)
            df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
            df = df.dropna(how='all').reset_index(drop=True)

            # Also read cell background colors to restore _status
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Find data start row (1-indexed, after header)
            xl_header_row = header_row + 2  # +1 for 0-index, +1 for header itself
            status_map = {}
            for xl_row in ws.iter_rows(min_row=xl_header_row):
                bg = xl_row[1].fill.fgColor.rgb if xl_row[1].fill and xl_row[1].fill.fgColor else 'FF000000'
                df_idx = xl_row[0].row - xl_header_row
                if 0 <= df_idx < len(df):
                    if bg in ('FFFFF176', 'FFFFF100', 'FFFFFF00', 'FFF0F076',
                              'FFFFF9C4', 'FFFEF9C3'):
                        status_map[df_idx] = STATUS_CHANGED
                    elif bg in ('FFD3D3D3', 'FFC0C0C0', 'FFD0D0D0', 'FFBDBDBD',
                                'FFE0E0E0', 'FFE8E8E8'):
                        status_map[df_idx] = STATUS_SOLD

            # Ensure internal columns
            df['_status'] = df.index.map(lambda i: status_map.get(i, STATUS_NONE))
            df['_change'] = ''
            if 'Observations' not in df.columns:
                df['Observations'] = ''

            # Auto-detect rows without price → STATUS_NO_PRICE (vectorized)
            if 'Price' in df.columns:
                empty_price = df['Price'].astype(str).str.strip().isin(
                    ('', 'nan', 'None', '-', 'N/A', 'n/a'))
                no_status   = df['_status'] == STATUS_NONE
                df.loc[empty_price & no_status, '_status'] = STATUS_NO_PRICE

            self.df = df
            self.current_file = path
            self.changes = []
            self._rebuild_search_cache()

            self._refresh_table()
            self._refresh_promotor_list()
            self._update_stats()
            self._status_var.set(f'Cargado: {Path(path).name} — {len(df)} propiedades')
            self.root.title(f'{APP_TITLE} · {Path(path).name}')

        except Exception as e:
            messagebox.showerror('Error al cargar Excel', str(e))

    # ── Table Rendering ──────────────────────────────────────────

    def _refresh_table(self, data: pd.DataFrame | None = None):
        self._tree.delete(*self._tree.get_children())
        src = data if data is not None else self.df

        # Sort: sold rows always go to the bottom; rest keep original order
        if not src.empty and '_status' in src.columns:
            src = src.iloc[src['_status'].map(
                lambda s: STATUS_ORDER.get(str(s), 0)
            ).argsort(kind='stable')]

        for row_num, (idx, row) in enumerate(src.iterrows()):
            values = []
            for col in MASTER_COLUMNS:
                v = row.get(col, '')
                values.append('' if (v is None or str(v) == 'nan') else str(v))

            st = str(row.get('_status', ''))
            if st == STATUS_CHANGED:
                tag = 'changed'
            elif st == STATUS_SOLD:
                tag = 'sold'
            elif st == STATUS_NO_PRICE:
                tag = 'no_price'
            else:
                # Zebra striping para filas sin estado especial
                tag = 'stripe' if row_num % 2 == 1 else 'normal'
            self._tree.insert('', tk.END, iid=str(idx), values=values, tags=(tag,))

        n = len(src)
        self._count_var.set(f'{n} propiedad{"es" if n != 1 else ""}')

    def _refresh_promotor_list(self):
        self._promotor_lb.delete(0, tk.END)
        if self.df.empty or 'Promotor' not in self.df.columns:
            return
        for p in sorted(self.df['Promotor'].dropna().unique()):
            self._promotor_lb.insert(tk.END, p)

    def _on_promotor_select(self, _event=None):
        sel = self._promotor_lb.curselection()
        if sel:
            self._apply_filters(promotor=self._promotor_lb.get(sel[0]))

    def _rebuild_search_cache(self):
        """Precalcula una columna con toda la fila concatenada en minúsculas."""
        if self.df.empty:
            self._search_cache = pd.Series(dtype=str)
            return
        self._search_cache = self.df[MASTER_COLUMNS].fillna('').astype(str).apply(
            lambda r: ' '.join(r).lower(), axis=1)

    def _on_search(self):
        if not hasattr(self, '_tree'):
            return   # toolbar se construye antes que el body; ignorar disparo prematuro
        q = self._search_var.get().lower()
        ph = 'buscar…'
        if q and q != ph:
            if not hasattr(self, '_search_cache') or len(self._search_cache) != len(self.df):
                self._rebuild_search_cache()
            self._refresh_table(self.df[self._search_cache.str.contains(q, regex=False)])
        else:
            self._refresh_table()

    def _apply_filters(self, promotor: str | None = None, status: str | None = None):
        data = self.df
        if promotor:
            data = data[data.get('Promotor', pd.Series(dtype=str)) == promotor]
        if status:
            data = data[data.get('_status', pd.Series(dtype=str)) == status]
        self._refresh_table(data)

    def _toggle_sort(self, col: str):
        if col not in self.df.columns:
            return
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self.df = self.df.sort_values(col, ascending=self._sort_asc,
                                       key=lambda x: x.fillna('').astype(str))
        self._refresh_table()

    def _update_stats(self):
        if self.df.empty:
            return
        total     = len(self.df)
        s         = self.df.get('_status', pd.Series(dtype=str))
        changed   = (s == STATUS_CHANGED).sum()
        sold      = (s == STATUS_SOLD).sum()
        no_price  = (s == STATUS_NO_PRICE).sum()
        promotors = self.df['Promotor'].nunique() if 'Promotor' in self.df.columns else 0
        txt = f'Total: {total}\nPromotoras: {promotors}'
        if changed:   txt += f'\nAmarillo: {changed}'
        if no_price:  txt += f'\nSin precio: {no_price}'
        if sold:      txt += f'\nAgotadas: {sold}'
        self._stats_lbl.config(text=txt)

    # ── CRUD ─────────────────────────────────────────────────────

    def cmd_add(self):
        self._open_editor(None)

    def cmd_edit(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo('Editar', 'Selecciona una fila para editar.')
            return
        self._open_editor(int(sel[0]))

    def cmd_delete(self):
        sel = self._tree.selection()
        if not sel:
            return
        if not messagebox.askyesno('Eliminar',
                '¿Eliminar la propiedad seleccionada?\nPuedes deshacer con Ctrl+Z.'):
            return
        self._snapshot()
        idx = int(sel[0])
        self.df = self.df.drop(idx).reset_index(drop=True)
        self._refresh_table()
        self._refresh_promotor_list()
        self._update_stats()

    def _open_editor(self, idx: int | None):
        dlg = tk.Toplevel(self.root)
        dlg.title('Nueva propiedad' if idx is None else 'Editar propiedad')
        dlg.geometry('520x580')
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(bg='#f5f7fa')

        tk.Label(dlg, text='Nueva propiedad' if idx is None else 'Editar propiedad',
                 font=('Arial', 13, 'bold'), bg='#f5f7fa', fg='#2c3e50'
                 ).grid(row=0, column=0, columnspan=2, pady=(12, 10), padx=12, sticky='w')

        row_data = {} if idx is None else self.df.loc[idx].to_dict()
        entries: dict[str, tk.StringVar] = {}

        edit_fields = MASTER_COLUMNS  # All visible columns
        for i, field in enumerate(edit_fields, 1):
            tk.Label(dlg, text=f'{field}:', font=('Arial', 10), bg='#f5f7fa',
                     anchor='e').grid(row=i, column=0, sticky='e', padx=(12, 6), pady=4)
            v = str(row_data.get(field, '') or '').replace('nan', '')
            var = tk.StringVar(value=v)
            ent = tk.Entry(dlg, textvariable=var, width=34, font=('Arial', 10),
                           relief='solid', bd=1)
            ent.grid(row=i, column=1, sticky='w', padx=(0, 12), pady=4)
            entries[field] = var

        # Status selector
        r = len(edit_fields) + 1
        tk.Label(dlg, text='Estado fila:', font=('Arial', 10), bg='#f5f7fa',
                 anchor='e').grid(row=r, column=0, sticky='e', padx=(12, 6), pady=4)
        st_var = tk.StringVar(value=str(row_data.get('_status', '') or ''))
        st_cb = ttk.Combobox(dlg, textvariable=st_var,
                              values=['', 'changed (amarillo)', 'sold (gris)'],
                              state='readonly', width=20, font=('Arial', 10))
        st_cb.grid(row=r, column=1, sticky='w', padx=(0, 12), pady=4)

        def save():
            self._snapshot()
            if idx is None:
                new = {f: entries[f].get() for f in edit_fields}
                new['_status'] = st_var.get().split(' ')[0] if st_var.get() else ''
                new['_change'] = ''
                self.df = pd.concat([self.df, pd.DataFrame([new])], ignore_index=True)
            else:
                for f in edit_fields:
                    self.df.at[idx, f] = entries[f].get()
                st = st_var.get().split(' ')[0] if st_var.get() else ''
                self.df.at[idx, '_status'] = st
            self._refresh_table()
            self._refresh_promotor_list()
            self._update_stats()
            dlg.destroy()

        bf = tk.Frame(dlg, bg='#f5f7fa')
        bf.grid(row=r+1, column=0, columnspan=2, pady=14)
        tk.Button(bf, text='💾  Guardar', command=save,
                  bg='#27ae60', fg='white', font=('Arial', 10, 'bold'),
                  padx=20, pady=6, relief='flat', cursor='hand2').pack(side=tk.LEFT, padx=6)
        tk.Button(bf, text='Cancelar', command=dlg.destroy,
                  bg='#95a5a6', fg='white', font=('Arial', 10),
                  padx=20, pady=6, relief='flat', cursor='hand2').pack(side=tk.LEFT, padx=6)

    # ── PDF Import ───────────────────────────────────────────────

    def cmd_import_pdfs(self):
        if not PDF_SUPPORT:
            messagebox.showerror('Librería no disponible',
                'Se necesita pdfplumber.\n\nInstala con:\n  pip install pdfplumber')
            return
        if self.df.empty:
            if not messagebox.askyesno('Sin datos',
                    'No hay listado maestro cargado.\n¿Cargar Excel maestro primero?'):
                return
            self.cmd_load_excel()
            if self.df.empty:
                return

        paths = filedialog.askopenfilenames(
            title='Seleccionar PDFs de promotoras',
            initialdir=str(SCRIPT_DIR),
            filetypes=[('PDF', '*.pdf'), ('Todos', '*.*')])
        if not paths:
            return

        self._show_pdf_preview(list(paths))

    def _show_pdf_preview(self, pdf_paths: list[str]):
        dlg = tk.Toplevel(self.root)
        dlg.title('Importar PDFs — previsualización de datos extraídos')
        dlg.geometry('780x560')
        dlg.grab_set()
        dlg.configure(bg='#f5f7fa')

        tk.Label(dlg, text='Datos extraídos de los PDFs',
                 font=('Arial', 13, 'bold'), bg='#f5f7fa', fg='#2c3e50'
                 ).pack(anchor='w', padx=14, pady=(12, 4))

        txt_frame = tk.Frame(dlg, bg='#f5f7fa')
        txt_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)

        txt = tk.Text(txt_frame, font=('Courier', 10), wrap='word',
                      bg='#1e1e1e', fg='#dcdcdc', relief='flat', bd=1,
                      insertbackground='white')
        sb = ttk.Scrollbar(txt_frame, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        txt.pack(fill=tk.BOTH, expand=True)

        # Color tags for the preview text
        txt.tag_configure('header', foreground='#61dafb', font=('Courier', 10, 'bold'))
        txt.tag_configure('ok',     foreground='#98c379')
        txt.tag_configure('warn',   foreground='#e5c07b')
        txt.tag_configure('err',    foreground='#e06c75')
        txt.tag_configure('unit',   foreground='#abb2bf')

        results: list[dict] = []

        for pdf_path in pdf_paths:
            name = Path(pdf_path).name
            txt.insert(tk.END, f'\n{"─"*62}\n', 'header')
            txt.insert(tk.END, f'📄  {name}\n', 'header')
            txt.insert(tk.END, f'{"─"*62}\n', 'header')

            try:
                units = extract_units_from_pdf(pdf_path)
                if units:
                    txt.insert(tk.END, f'✅  {len(units)} unidades detectadas\n\n', 'ok')
                    for u in units[:8]:
                        price_str = f"{u.get('price', 0):,.0f} €" if u.get('price') else '—'
                        line = (f"  • Precio: {price_str:<14}  "
                                f"Dorm: {u.get('bedrooms','?'):<4}  "
                                f"Planta: {str(u.get('floor','?')):<18}  "
                                f"Estado: {u.get('status','?')}\n")
                        txt.insert(tk.END, line, 'unit')
                    if len(units) > 8:
                        txt.insert(tk.END, f'  … y {len(units)-8} unidades más\n', 'warn')
                    results.append({'file': name, 'path': pdf_path, 'data': units})
                else:
                    txt.insert(tk.END,
                        '⚠️  No se pudieron extraer datos estructurados.\n'
                        '   Posiblemente el PDF es una imagen (requiere OCR).\n'
                        '   → Este PDF quedará marcado para revisión manual.\n', 'warn')
                    results.append({'file': name, 'path': pdf_path, 'data': [], 'manual': True})
            except Exception as e:
                txt.insert(tk.END, f'❌  Error: {e}\n', 'err')
                results.append({'file': name, 'path': pdf_path, 'data': [], 'error': str(e)})

        txt.insert(tk.END,
            f'\n{"─"*62}\n'
            '¿Deseas aplicar los cambios detectados al listado maestro?\n'
            '(Solo se modificarán filas donde se encuentre un cambio confirmado)\n',
            'warn')
        txt.configure(state='disabled')

        bf = tk.Frame(dlg, bg='#f5f7fa')
        bf.pack(pady=10)

        def apply():
            dlg.destroy()
            self._apply_pdf_changes(results)

        tk.Button(bf, text='✅  Aplicar cambios', command=apply,
                  bg='#27ae60', fg='white', font=('Arial', 11, 'bold'),
                  padx=18, pady=7, relief='flat', cursor='hand2').pack(side=tk.LEFT, padx=6)
        tk.Button(bf, text='❌  Cancelar', command=dlg.destroy,
                  bg='#c0392b', fg='white', font=('Arial', 11),
                  padx=18, pady=7, relief='flat', cursor='hand2').pack(side=tk.LEFT, padx=6)

    def _apply_pdf_changes(self, results: list[dict]):
        n_substitutions = 0
        n_sold_out      = 0
        manual_review   = []

        self._snapshot()
        for result in results:
            if result.get('manual') or result.get('error') or not result.get('data'):
                manual_review.append(result['file'])
                continue

            pdf_units: list[dict] = result['data']
            all_pdf_prices   = {u.get('price') for u in pdf_units}
            avail_prices     = {u['price'] for u in pdf_units
                                if u.get('status') == 'AVAILABLE' and u.get('price')}

            for idx, row in self.df.iterrows():
                master_price = parse_price(row.get('Price'))
                if master_price is None:
                    continue  # Skip non-numeric prices (SOLD, etc.)

                price_in_pdf = master_price in all_pdf_prices

                if not price_in_pdf:
                    # The unit may have been sold or price changed
                    bedrooms = row.get('Bedrooms')
                    try:
                        bedrooms = int(bedrooms) if bedrooms and str(bedrooms) != 'nan' else None
                    except (ValueError, TypeError):
                        bedrooms = None

                    floor_cat = classify_floor(str(row.get('Floor', '')))

                    # Look for replacement in same category
                    replacement = self._find_replacement(
                        pdf_units, bedrooms, floor_cat, master_price)

                    if replacement:
                        old_price = master_price
                        new_price = replacement['price']
                        self.df.at[idx, 'Price'] = str(int(new_price)) if new_price == int(new_price) else str(new_price)
                        self.df.at[idx, '_status'] = STATUS_CHANGED
                        change_desc = (f'Precio: {old_price:,.0f}€ → {new_price:,.0f}€  '
                                       f'(Δ {new_price - old_price:+,.0f}€)')
                        self.df.at[idx, '_change'] = change_desc
                        self.changes.append({
                            'type': 'price_change',
                            'promotor': row.get('Promotor', ''),
                            'development': row.get('Development', ''),
                            'reference': row.get('Reference', ''),
                            'old_price': old_price,
                            'new_price': new_price,
                            'file': result['file'],
                            'idx': idx,
                        })
                        n_substitutions += 1

                    else:
                        # No replacement found → category sold out
                        self.df.at[idx, '_status'] = STATUS_SOLD
                        self.df.at[idx, 'Observations'] = 'Categoría agotada — archivar en Odoo'
                        self.df.at[idx, '_change'] = 'Sin reemplazo disponible en el PDF'
                        self.changes.append({
                            'type': 'sold_out',
                            'promotor': row.get('Promotor', ''),
                            'development': row.get('Development', ''),
                            'reference': row.get('Reference', ''),
                            'price': master_price,
                            'file': result['file'],
                            'idx': idx,
                        })
                        n_sold_out += 1

        self._refresh_table()
        self._update_stats()

        summary_lines = ['Procesamiento completado:\n']
        if n_substitutions:
            summary_lines.append(f'🟡  {n_substitutions} cambios de precio / sustituciones (filas amarillas)')
        if n_sold_out:
            summary_lines.append(f'⬜  {n_sold_out} categorías agotadas (filas grises)')
        if not n_substitutions and not n_sold_out:
            summary_lines.append('✅  Sin cambios detectados (todos los precios coinciden)')
        if manual_review:
            summary_lines.append(f'\n⚠️  {len(manual_review)} PDF(s) requieren revisión manual:')
            for f in manual_review:
                summary_lines.append(f'    • {f}')

        messagebox.showinfo('Cambios aplicados', '\n'.join(summary_lines))
        self._status_var.set(
            f'PDFs procesados — {n_substitutions} cambios, {n_sold_out} agotadas')

    # ── Matching engine ─────────────────────────────────────────────────────────

    # Umbral máximo de variación de precio antes de etiquetar como sospechoso.
    # Cambios superiores al 40 % se marcan para revisión manual en la previsualización
    # pero SÍ se aplican — el usuario puede deshacer con Ctrl+Z.
    PRICE_CHANGE_WARN_PCT = 0.40

    @staticmethod
    def _pdf_has_refs(pdf_units: list[dict]) -> bool:
        """True si al menos el 30 % de las unidades del PDF tienen referencia."""
        if not pdf_units:
            return False
        n_with_ref = sum(1 for u in pdf_units if str(u.get('reference', '')).strip())
        return n_with_ref >= max(1, len(pdf_units) * 0.30)

    @staticmethod
    def _find_best(pdf_units: list[dict], bedrooms: int | None,
                   floor_cat: str, used_ids: set) -> dict | None:
        """
        Devuelve la unidad disponible más barata que coincida en dormitorios y planta,
        excluyendo las ya asignadas a otras filas del Excel (used_ids).
        """
        avail = [u for u in pdf_units
                 if u.get('status') == 'AVAILABLE' and u.get('price')
                 and id(u) not in used_ids]
        if not avail:
            return None

        floor_cats    = {id(u): classify_floor(str(u.get('floor', ''))) for u in avail}
        pdf_has_floor = any(fc != 'other' for fc in floor_cats.values())

        candidates = []
        for u in avail:
            u_beds  = u.get('bedrooms')
            u_floor = floor_cats[id(u)]
            bed_ok  = bedrooms is None or u_beds is None or int(u_beds) == int(bedrooms)
            floor_ok = (not pdf_has_floor or floor_cat == 'other'
                        or u_floor == floor_cat)
            if bed_ok and floor_ok:
                candidates.append(u)

        return min(candidates, key=lambda x: x.get('price', float('inf'))) if candidates else None

    @staticmethod
    def _match_row_to_pdf(
        master_price: float,
        master_ref:   str,
        beds_int:     'int | None',
        floor_cat:    str,
        pdf_units:    list[dict],
        used_ids:     set,
    ) -> 'tuple[str, dict | None]':
        """
        Compara una fila del Excel maestro con las unidades del PDF.

        Estrategia 1 — matching por REFERENCIA (cuando el PDF incluye refs):
          · Referencia encontrada, precio igual  → 'no_change'
          · Referencia encontrada, precio distinto → 'price_change'  (misma unidad)
          · Referencia no encontrada → unidad vendida → buscar sustituto
              · Sustituto encontrado → 'replacement'
              · Sin sustituto       → 'sold_out'

        Estrategia 2 — matching por PRECIO (PDFs sin columna de referencia):
          · Precio encontrado en disponibles → 'no_change'
          · No encontrado → buscar sustituto → 'replacement' o 'sold_out'

        Retorna (accion, unidad_pdf):
          accion: 'no_change' | 'price_change' | 'replacement' | 'sold_out'
        """
        # ── Estrategia 1: por referencia ──────────────────────────────────────
        if PropertyManagerApp._pdf_has_refs(pdf_units) and master_ref.strip():
            ref_norm = master_ref.strip().upper()
            ref_unit = next(
                (u for u in pdf_units
                 if str(u.get('reference', '')).strip().upper() == ref_norm
                 and id(u) not in used_ids),
                None,
            )
            if ref_unit is not None:
                used_ids.add(id(ref_unit))
                if abs(ref_unit.get('price', 0) - master_price) < 1:
                    return 'no_change', ref_unit
                return 'price_change', ref_unit

            # Referencia no encontrada → vendida → buscar sustituto
            subst = PropertyManagerApp._find_best(pdf_units, beds_int, floor_cat, used_ids)
            if subst:
                used_ids.add(id(subst))
                return 'replacement', subst
            return 'sold_out', None

        # ── Estrategia 2: por precio ──────────────────────────────────────────
        avail_prices = {u['price'] for u in pdf_units
                        if u.get('status') == 'AVAILABLE' and u.get('price')
                        and id(u) not in used_ids}
        if master_price in avail_prices:
            return 'no_change', None

        subst = PropertyManagerApp._find_best(pdf_units, beds_int, floor_cat, used_ids)
        if subst:
            used_ids.add(id(subst))
            return 'replacement', subst
        return 'sold_out', None

    # ── Legacy wrapper (usado internamente) ──────────────────────────────────

    @staticmethod
    def _find_replacement(pdf_units: list[dict], bedrooms: int | None,
                           floor_cat: str, exclude_price: float) -> dict | None:
        """Mantiene compatibilidad con llamadas directas antiguas."""
        used: set = set()
        return PropertyManagerApp._find_best(pdf_units, bedrooms, floor_cat, used)

    # ── Auto-Update from folder ──────────────────────────────────

    def cmd_auto_update(self):
        """Escanea la carpeta en busca de PDFs de precios y actualiza el Excel maestro."""
        if not PDF_SUPPORT:
            messagebox.showerror('Librería no disponible',
                'Se necesita pdfplumber.\n\nInstala con:\n  pip install pdfplumber')
            return

        if self.df.empty:
            self._auto_load()
            if self.df.empty:
                messagebox.showinfo('Sin datos',
                    'Carga primero el Excel maestro para poder actualizar.')
                return

        folder = LISTADOS_DIR if LISTADOS_DIR.exists() else (
            Path(self.current_file).parent if self.current_file else SCRIPT_DIR
        )
        all_pdfs   = sorted(folder.glob('*.pdf'))
        price_pdfs = [p for p in all_pdfs if es_listado_precios(p.stem)]

        if not price_pdfs:
            messagebox.showinfo('Sin PDFs de precios',
                f'No se encontraron PDFs de listados de precios en:\n{folder}\n\n'
                f'PDFs totales en carpeta: {len(all_pdfs)}\n\n'
                'Los PDFs deben tener en su nombre palabras como:\n'
                '"Price List", "Lista de Precios", "Pricelist", "Precios"…')
            return

        self._show_update_dialog(price_pdfs, folder)

    def _match_pdf_to_dev(self, pdf_path: Path,
                          exclude_devs: set[str] | None = None
                          ) -> tuple[str | None, str, float]:
        """
        Intenta asociar un PDF a un desarrollo del Excel maestro.
        Devuelve (nombre_desarrollo, metodo, score).
        exclude_devs: conjunto de desarrollos ya asignados a otro PDF (se ignoran).
        Puntuación = suma de longitud de chars de palabras del desarrollo que coinciden
        en el texto del PDF (con tolerancia a un typo para palabras largas).
        """
        if self.df.empty or 'Development' not in self.df.columns:
            return None, 'none', 0.0

        developments = [d for d in self.df['Development'].dropna().unique()
                        if not exclude_devs or d not in exclude_devs]
        filename_norm = _normalize_for_match(pdf_path.stem)

        best_dev   = None
        best_score = 0.0

        # 1. Coincidencia por nombre de archivo
        for dev in developments:
            score = _dev_score_in_text(dev, filename_norm)
            if score > best_score:
                best_score = score
                best_dev   = dev

        if best_dev and best_score > 0:
            return best_dev, 'filename', best_score

        # 2. Coincidencia por contenido del PDF (texto digital)
        try:
            with pdfplumber.open(str(pdf_path)) as pdf:
                content = _normalize_for_match(
                    ' '.join(page.extract_text() or '' for page in pdf.pages[:3]))
            if content.strip():
                for dev in developments:
                    score = _dev_score_in_text(dev, content)
                    if score > best_score:
                        best_score = score
                        best_dev   = dev
                if best_dev and best_score > 0:
                    return best_dev, 'content', best_score
        except Exception:
            pass

        # 3. Coincidencia por OCR de la primera página (PDFs de imagen)
        if OCR_SUPPORT:
            try:
                doc = fitz.open(str(pdf_path))
                page = doc[0]
                mat = fitz.Matrix(1.5, 1.5)   # resolución moderada para matching
                pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
                img = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)
                doc.close()
                ocr_content = _normalize_for_match(
                    pytesseract.image_to_string(img, config='--psm 11 --oem 3'))
                if ocr_content.strip():
                    for dev in developments:
                        score = _dev_score_in_text(dev, ocr_content)
                        if score > best_score:
                            best_score = score
                            best_dev   = dev
                    if best_dev and best_score > 0:
                        return best_dev, 'ocr', best_score
            except Exception:
                pass

        return None, 'none', 0.0

    def _show_update_dialog(self, price_pdfs: list[Path], folder: Path):
        """Ventana de progreso y resultados de la actualización automática."""
        import threading

        dlg = tk.Toplevel(self.root)
        dlg.title('Actualizar Listados')
        dlg.geometry('960x680')
        dlg.transient(self.root)   # queda encima de la ventana principal
        # NO grab_set() — permite abrir la ventana de detalle en paralelo
        dlg.configure(bg=BG_APP)

        # ── Cabecera ──────────────────────────────────────────────
        hdr = tk.Frame(dlg, bg=BG_TOOLBAR)
        hdr.pack(fill=tk.X)
        tk.Frame(hdr, bg=C_SUCCESS, height=3).pack(fill=tk.X)
        hdr_inner = tk.Frame(hdr, bg=BG_TOOLBAR, pady=10)
        hdr_inner.pack(fill=tk.X, padx=14)
        tk.Label(hdr_inner, text='⟳  Actualizar Listados de Precios',
                 font=(FONT_UI, 13, 'bold'), bg=BG_TOOLBAR, fg='white').pack(anchor='w')
        tk.Label(hdr_inner, text=f'📁  {folder}   ·   {len(price_pdfs)} PDFs encontrados',
                 font=(FONT_UI, 9), bg=BG_TOOLBAR, fg='#64748B').pack(anchor='w', pady=(2, 0))

        # ── Tabla de resultados ───────────────────────────────────
        tree_frame = tk.Frame(dlg, bg=BG_APP)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 0))

        cols = ('pdf', 'desarrollo', 'unidades', 'resultado')
        tree = ttk.Treeview(tree_frame, columns=cols, show='headings', height=10)
        tree.heading('pdf',        text='  PDF')
        tree.heading('desarrollo', text='  Desarrollo')
        tree.heading('unidades',   text='  Unidades')
        tree.heading('resultado',  text='  Resultado')
        tree.column('pdf',        width=300, stretch=True)
        tree.column('desarrollo', width=200, stretch=True)
        tree.column('unidades',   width=80,  anchor='center')
        tree.column('resultado',  width=290, stretch=True)

        tree.tag_configure('ok',      background='#ECFDF5', foreground='#065F46')
        tree.tag_configure('changed', background=COLOR_YELLOW, foreground='#92400E')
        tree.tag_configure('nomatch', background='#FEF2F2',   foreground='#991B1B')
        tree.tag_configure('manual',  background='#FFFBEB',   foreground='#92400E')

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ── Log detallado ──────────────────────────────────────────
        log_frame = tk.Frame(dlg, bg=BG_APP)
        log_frame.pack(fill=tk.X, padx=10, pady=(6, 0))
        tk.Label(log_frame, text='Log de proceso', font=(FONT_UI, 8, 'bold'),
                 bg=BG_APP, fg=T_MUTED, anchor='w').pack(fill=tk.X, pady=(0, 2))
        log_txt = tk.Text(log_frame, height=8,
                          font=(FONT_MONO, 9),
                          bg='#0D1117', fg='#C9D1D9',
                          relief='flat', state='disabled',
                          padx=10, pady=8,
                          highlightthickness=0)
        log_sb = ttk.Scrollbar(log_frame, command=log_txt.yview)
        log_txt.configure(yscrollcommand=log_sb.set)
        log_sb.pack(side=tk.RIGHT, fill=tk.Y)
        log_txt.pack(fill=tk.X)
        log_txt.tag_configure('ok',   foreground='#3FB950')  # green
        log_txt.tag_configure('warn', foreground='#D29922')  # amber
        log_txt.tag_configure('err',  foreground='#F85149')  # red
        log_txt.tag_configure('info', foreground='#58A6FF')  # blue

        def log(msg, tag='info'):
            log_txt.configure(state='normal')
            log_txt.insert(tk.END, msg + '\n', tag)
            log_txt.see(tk.END)
            log_txt.configure(state='disabled')
            dlg.update_idletasks()

        # ── Botones ───────────────────────────────────────────────
        btn_frame = tk.Frame(dlg, bg=BG_TOOLBAR, pady=10)
        btn_frame.pack(fill=tk.X)
        tk.Frame(btn_frame, bg='#334155', height=1).pack(fill=tk.X)
        btn_inner = tk.Frame(btn_frame, bg=BG_TOOLBAR)
        btn_inner.pack(pady=8, padx=14)

        def _dbtn(text, cmd, color, state='normal'):
            b = tk.Button(btn_inner, text=text, command=cmd,
                          bg=color, fg='white',
                          font=(FONT_UI, 10, 'bold'),
                          padx=18, pady=7, relief='flat', cursor='hand2',
                          state=state, bd=0,
                          activebackground=color, activeforeground='white')
            b.pack(side=tk.LEFT, padx=4)
            return b

        apply_btn   = _dbtn('✓  Aplicar cambios',       lambda: None, C_SUCCESS,  'disabled')
        preview_btn = _dbtn('⊞  Ver cambios detallados', lambda: None, C_INFO,     'disabled')
        _dbtn('✕  Cerrar', dlg.destroy, C_DANGER)

        matched_results: list[dict] = []

        # ── Procesamiento en hilo de fondo ────────────────────────
        # Mantiene la UI viva (barra de progreso y log se actualizan en tiempo real).

        def _safe_log(msg: str, tag: str = 'info'):
            def _do():
                log_txt.configure(state='normal')
                log_txt.insert(tk.END, msg + '\n', tag)
                log_txt.see(tk.END)
                log_txt.configure(state='disabled')
            try:
                dlg.after(0, _do)
            except Exception:
                pass

        def _safe_tree(values, tags):
            try:
                dlg.after(0, lambda v=values, t=tags: tree.insert('', tk.END, values=v, tags=(t,)))
            except Exception:
                pass

        def _process():
            _safe_log(f'Escaneando {len(price_pdfs)} PDFs en {folder}…', 'info')
            if OCR_SUPPORT:
                _safe_log('✅ OCR activo (Tesseract instalado y funcional)', 'ok')
            else:
                _safe_log('⚠ OCR no disponible — solo PDFs con texto digital', 'warn')
                if OCR_ERROR:
                    _safe_log(f'   Motivo: {OCR_ERROR}', 'warn')

            developments = (self.df['Development'].dropna().unique()
                            if 'Development' in self.df.columns else [])
            dev_pdfs:  dict[str, list[tuple[Path, str]]] = {}
            unmatched: list[Path] = []

            # ── Fase 1: asociar PDFs a desarrollos ───────────────
            for pdf_path in price_pdfs:
                fn_norm    = _normalize_for_match(pdf_path.stem)
                scores     = {dev: _dev_score_in_text(dev, fn_norm) for dev in developments}
                best_score = max(scores.values(), default=0)

                if best_score == 0:
                    try:
                        with pdfplumber.open(str(pdf_path)) as _p:
                            content = _normalize_for_match(
                                ' '.join(pg.extract_text() or '' for pg in _p.pages[:3]))
                        if content.strip():
                            scores     = {dev: _dev_score_in_text(dev, content)
                                          for dev in developments}
                            best_score = max(scores.values(), default=0)
                    except Exception:
                        pass

                if best_score == 0:
                    unmatched.append(pdf_path)
                    continue

                best_dev = max(scores, key=lambda d: scores[d])
                method   = ('filename' if _dev_score_in_text(best_dev, fn_norm) > 0
                            else 'content')
                dev_pdfs.setdefault(best_dev, []).append((pdf_path, method))

            # ── Fase 2: PDFs sin coincidencia ────────────────────
            for pdf_path in unmatched:
                sn = pdf_path.name[:55] + ('…' if len(pdf_path.name) > 55 else '')
                _safe_log(f'\n─── {pdf_path.name}', 'info')
                _safe_log(f'  ⚠ No se pudo asociar a ningún desarrollo del Excel', 'warn')
                _safe_tree((sn, '—', '—', '⚠ Sin coincidencia en Excel'), 'nomatch')
                matched_results.append({'pdf': pdf_path, 'dev': None,
                                        'units': [], 'tag': 'nomatch'})

            # ── Fase 3: procesar desarrollo por desarrollo ────────
            for dev_name, pdfs_for_dev in dev_pdfs.items():
                pdf_names = [p.name[:40] for p, _ in pdfs_for_dev]
                _safe_log(f'\n─── Desarrollo: "{dev_name}" ({len(pdfs_for_dev)} PDF(s))', 'info')
                for p, method in pdfs_for_dev:
                    _safe_log(f'  ↳ {p.name[:60]}  (por {method})', 'ok')

                all_units: list[dict] = []
                first_pdf = pdfs_for_dev[0][0]
                errors:   list[str]  = []

                for pdf_path, _ in pdfs_for_dev:
                    try:
                        units = extract_units_from_pdf(str(pdf_path))
                    except Exception as e:
                        errors.append(f'{pdf_path.name[:40]}: {e}')
                        continue

                    if units and units[0].get('_ocr_error'):
                        errors.append(f'{pdf_path.name[:40]}: {units[0]["_ocr_error"]}')
                        continue

                    if not units:
                        errors.append(f'{pdf_path.name[:40]}: sin unidades extraíbles')
                        continue

                    all_units.extend(units)
                    avail   = [u for u in units if u.get('status') == 'AVAILABLE']
                    ocr_tag = ' (OCR)' if any(u.get('ocr') for u in units) else ''
                    _safe_log(f'  → {len(units)} unidades{ocr_tag} ({len(avail)} disponibles)', 'ok')

                for err in errors:
                    _safe_log(f'  ⚠ {err}', 'warn')

                pdf_label = ' + '.join(n[:35] for n in pdf_names)

                if not all_units:
                    _safe_tree((pdf_label, dev_name, 0, '⚠ Sin datos — revisión manual'),
                               'manual')
                    matched_results.append({'pdf': first_pdf, 'dev': dev_name,
                                            'units': [], 'tag': 'manual'})
                    continue

                # Deduplicar unidades fusionadas
                seen_keys: set = set()
                merged:    list[dict] = []
                for u in all_units:
                    key = (round(u.get('price', 0)), u.get('status', ''),
                           u.get('bedrooms'), str(u.get('floor', ''))[:10],
                           str(u.get('reference', ''))[:15])
                    if key not in seen_keys:
                        seen_keys.add(key)
                        merged.append(u)
                all_units = merged

                n_changes    = self._count_changes(dev_name, all_units)
                n_unverified = self._count_ocr_unverified(dev_name, all_units)
                if n_changes > 0:
                    result_txt = f'🟡 {n_changes} cambio(s) detectado(s)'
                    tag_r = 'changed'
                    _safe_log(f'  → {result_txt}', 'warn')
                elif n_unverified > 0:
                    result_txt = f'✅ Sin cambios ({n_unverified} sin verificar — OCR)'
                    tag_r = 'ok'
                    _safe_log(f'  → {result_txt}', 'ok')
                else:
                    result_txt = '✅ Sin cambios confirmados'
                    tag_r = 'ok'
                    _safe_log(f'  → {result_txt}', 'ok')

                _safe_tree((pdf_label, dev_name, len(all_units), result_txt), tag_r)
                matched_results.append({'pdf': first_pdf, 'dev': dev_name,
                                        'units': all_units, 'tag': tag_r,
                                        'n_changes': n_changes})

            _safe_log('\nAnálisis completado. Pulsa "Aplicar cambios" para actualizar el Excel.', 'ok')

            # Activar botones — debe hacerse en el hilo principal
            def _enable_buttons():
                has_data = any(r.get('units') for r in matched_results)
                if has_data:
                    def _open_preview():
                        prev = self._show_change_preview(matched_results)
                        if prev:
                            prev.lift()
                            prev.focus_set()

                    apply_btn.configure(
                        state='normal',
                        command=lambda: self._apply_auto_update(matched_results, dlg))
                    preview_btn.configure(state='normal', command=_open_preview)

            try:
                dlg.after(0, _enable_buttons)
            except Exception:
                pass

        threading.Thread(target=_process, daemon=True).start()

    def _show_change_preview(self, matched_results: list[dict]):
        """
        Muestra una ventana detallada con TODOS los cambios que se aplicarían,
        fila a fila: desarrollo, referencia, precio actual → precio nuevo / estado.
        """
        prev = tk.Toplevel(self.root)
        prev.title('🔍 Vista previa de cambios')
        prev.geometry('1020x640')
        prev.configure(bg='#f5f7fa')
        prev.lift()
        prev.focus_set()

        tk.Label(prev,
                 text='Vista previa — cambios que se aplicarán al Excel maestro',
                 font=('Arial', 12, 'bold'), bg='#f5f7fa', fg='#2c3e50'
                 ).pack(anchor='w', padx=14, pady=(10, 4))

        tk.Label(prev,
                 text='Revisa cada fila antes de pulsar "Aplicar cambios" en la ventana anterior.',
                 font=('Arial', 9), bg='#f5f7fa', fg='#7f8c8d'
                 ).pack(anchor='w', padx=14, pady=(0, 6))

        # ── Tabla de previsualización ─────────────────────────────
        cols = ('desarrollo', 'ref', 'dormitorios', 'planta',
                'precio_actual', 'precio_nuevo', 'accion', 'pdf')
        hdr_labels = ('Desarrollo', 'Referencia', 'Dorm.', 'Planta',
                      'Precio actual', 'Precio nuevo / PDF', 'Acción', 'PDF origen')

        frm = tk.Frame(prev, bg='#f5f7fa')
        frm.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        tv = ttk.Treeview(frm, columns=cols, show='headings', height=20)
        for col, lbl in zip(cols, hdr_labels):
            tv.heading(col, text=lbl)
        tv.column('desarrollo',   width=160)
        tv.column('ref',          width=110)
        tv.column('dormitorios',  width=55,  anchor='center')
        tv.column('planta',       width=110)
        tv.column('precio_actual',width=110, anchor='e')
        tv.column('precio_nuevo', width=160, anchor='e')
        tv.column('accion',       width=130, anchor='center')
        tv.column('pdf',          width=170)

        tv.tag_configure('price_change', background='#fff9c4', foreground='#7d6000')
        tv.tag_configure('sold_out',     background='#e0e0e0', foreground='#424242')
        tv.tag_configure('no_change',    background='#e8f5e9', foreground='#2e7d32')
        tv.tag_configure('manual',       background='#fff3e0', foreground='#e65100')

        vsb = ttk.Scrollbar(frm, orient=tk.VERTICAL, command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        tv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ── Leyenda ───────────────────────────────────────────────
        leg = tk.Frame(prev, bg='#f5f7fa')
        leg.pack(fill=tk.X, padx=10, pady=(2, 0))
        for color, texto in [('#fff9c4','Cambio de precio'),
                              ('#FFCCCC','Cambio >40 % — verificar'),
                              ('#e0e0e0','Sin reemplazo (categoría agotada)'),
                              ('#e8f5e9','Sin cambio — precio coincide'),
                              ('#fff3e0','Revisión manual')]:
            tk.Label(leg, text=f'  {texto}  ', bg=color, font=('Arial', 9),
                     relief='solid', bd=1).pack(side=tk.LEFT, padx=4, pady=3)

        tv.tag_configure('suspicious', background='#FFCCCC', foreground='#7B0000')

        # ── Contadores ────────────────────────────────────────────
        n_price = 0; n_sold = 0; n_ok = 0; n_manual = 0

        # ── Construir filas usando el motor de matching unificado ──
        for result in matched_results:
            dev_name  = result.get('dev')
            pdf_units = result.get('units', [])
            pdf_name  = result['pdf'].name[:40] if result.get('pdf') else '—'

            if not dev_name or not pdf_units:
                # PDF sin datos / sin coincidencia
                tv.insert('', tk.END,
                          values=('—', '—', '—', '—', '—', '—',
                                  '⚠ Revisión manual', pdf_name),
                          tags=('manual',))
                n_manual += 1
                continue

            is_ocr = any(u.get('ocr') for u in pdf_units)

            if 'Development' not in self.df.columns:
                continue
            dev_rows  = self.df[self.df['Development'] == dev_name]
            used_ids: set = set()   # evita asignar la misma unidad a dos filas

            for _, row in dev_rows.iterrows():
                master_price = parse_price(row.get('Price'))
                ref   = str(row.get('Reference', '') or '—')
                beds  = str(row.get('Bedrooms',  '') or '—')
                floor = str(row.get('Floor',     '') or '—')

                if master_price is None:
                    continue

                precio_actual_str = f'{master_price:,.0f} €'

                try:
                    beds_raw = row.get('Bedrooms')
                    beds_int = int(float(str(beds_raw))) if (
                        beds_raw is not None
                        and str(beds_raw) not in ('', 'nan', '—')
                    ) else None
                except (ValueError, TypeError):
                    beds_int = None
                floor_cat = classify_floor(floor)
                master_ref = str(row.get('Reference', '') or '')

                if is_ocr:
                    # ── PDF de imagen (OCR): matching por categoría ───────────
                    avail_ocr   = [u for u in pdf_units
                                   if u.get('status') == 'AVAILABLE' and u.get('price')]
                    beds_filter = (beds_int is None)
                    strict = [u for u in avail_ocr
                              if (beds_filter or u.get('bedrooms') == beds_int)
                              and u.get('floor')
                              and (classify_floor(u['floor']) == floor_cat
                                   or floor_cat == 'other')]
                    loose  = [u for u in avail_ocr
                              if (beds_filter or u.get('bedrooms') == beds_int)
                              and not u.get('floor')]

                    if strict:
                        if any(abs(u['price'] - master_price) < 1 for u in strict):
                            tv.insert('', tk.END,
                                      values=(dev_name, ref, beds, floor,
                                              precio_actual_str, '—',
                                              '✅ Sin cambio', pdf_name),
                                      tags=('no_change',)); n_ok += 1
                        elif any(abs(u['price'] - master_price) < 1 for u in loose):
                            tv.insert('', tk.END,
                                      values=(dev_name, ref, beds, floor,
                                              precio_actual_str, '—',
                                              '✅ Sin cambio', pdf_name),
                                      tags=('no_change',)); n_ok += 1
                        else:
                            best  = min(strict, key=lambda u: u['price'])
                            new_p = best['price']
                            delta = new_p - master_price
                            pct   = abs(delta) / master_price if master_price else 0
                            tag_u = 'suspicious' if pct > self.PRICE_CHANGE_WARN_PCT else 'price_change'
                            lbl   = '⚠ Verificar (>40%)' if pct > self.PRICE_CHANGE_WARN_PCT else '🟡 Precio cambiado (OCR)'
                            tv.insert('', tk.END,
                                      values=(dev_name, ref, beds, floor,
                                              precio_actual_str,
                                              f'{new_p:,.0f} €  ({delta:+,.0f} €)',
                                              lbl, pdf_name),
                                      tags=(tag_u,)); n_price += 1
                    elif any(abs(u['price'] - master_price) < 1 for u in loose):
                        tv.insert('', tk.END,
                                  values=(dev_name, ref, beds, floor,
                                          precio_actual_str, '—',
                                          '✅ Sin cambio', pdf_name),
                                  tags=('no_change',)); n_ok += 1
                    else:
                        tv.insert('', tk.END,
                                  values=(dev_name, ref, beds, floor,
                                          precio_actual_str, '—',
                                          '⚠ OCR: verificar manual', pdf_name),
                                  tags=('manual',)); n_manual += 1

                else:
                    # ── PDF de texto: matching por referencia (o precio si no hay refs) ──
                    action, pdf_unit = self._match_row_to_pdf(
                        master_price, master_ref, beds_int, floor_cat,
                        pdf_units, used_ids)

                    if action == 'no_change':
                        tv.insert('', tk.END,
                                  values=(dev_name, ref, beds, floor,
                                          precio_actual_str, '—',
                                          '✅ Sin cambio', pdf_name),
                                  tags=('no_change',)); n_ok += 1

                    elif action in ('price_change', 'replacement'):
                        new_p     = pdf_unit['price']
                        delta     = new_p - master_price
                        pct       = abs(delta) / master_price if master_price else 0
                        nuevo_str = f'{new_p:,.0f} €  ({delta:+,.0f} €)'
                        if pct > self.PRICE_CHANGE_WARN_PCT:
                            lbl   = '⚠ Verificar (>40%)'
                            tag_u = 'suspicious'
                        elif action == 'replacement':
                            lbl   = '🔄 Sustitución'
                            tag_u = 'price_change'
                        else:
                            lbl   = '🟡 Precio cambiado'
                            tag_u = 'price_change'
                        tv.insert('', tk.END,
                                  values=(dev_name, ref, beds, floor,
                                          precio_actual_str, nuevo_str,
                                          lbl, pdf_name),
                                  tags=(tag_u,)); n_price += 1

                    else:  # sold_out
                        tv.insert('', tk.END,
                                  values=(dev_name, ref, beds, floor,
                                          precio_actual_str, '—',
                                          '⬜ Categ. agotada', pdf_name),
                                  tags=('sold_out',)); n_sold += 1

        # ── Resumen ───────────────────────────────────────────────
        resumen = tk.Label(
            prev,
            text=(f'Resumen:  🟡 {n_price} cambios de precio   '
                  f'⬜ {n_sold} categorías agotadas   '
                  f'✅ {n_ok} sin cambio   '
                  f'⚠ {n_manual} revisión manual'),
            font=('Arial', 10, 'bold'), bg='#2c3e50', fg='white', pady=5)
        resumen.pack(fill=tk.X, padx=0, pady=(4, 0))

        tk.Button(prev, text='Cerrar vista previa', command=prev.destroy,
                  bg='#7f8c8d', fg='white', font=('Arial', 10),
                  padx=16, pady=6, relief='flat', cursor='hand2').pack(pady=8)

        return prev

    def _count_ocr_unverified(self, dev_name: str, pdf_units: list[dict]) -> int:
        """Cuenta cuántas filas quedarían como 'OCR: verificar manual' (solo PDFs de imagen)."""
        if 'Development' not in self.df.columns: return 0
        if not any(u.get('ocr') for u in pdf_units): return 0
        dev_rows  = self.df[self.df['Development'] == dev_name]
        avail_ocr = [u for u in pdf_units if u.get('status') == 'AVAILABLE' and u.get('price')]
        n = 0
        for _, row in dev_rows.iterrows():
            master_price = parse_price(row.get('Price'))
            if master_price is None: continue
            try:
                beds_raw = row.get('Bedrooms')
                beds_int = int(float(str(beds_raw))) if (
                    beds_raw is not None and str(beds_raw) not in ('', 'nan', '—')
                ) else None
            except (ValueError, TypeError):
                beds_int = None
            floor_cat   = classify_floor(str(row.get('Floor', '')))
            beds_filter = (beds_int is None)
            strict = [u for u in avail_ocr
                      if (beds_filter or u.get('bedrooms') == beds_int)
                      and u.get('floor')
                      and (classify_floor(u['floor']) == floor_cat or floor_cat == 'other')]
            loose  = [u for u in avail_ocr
                      if (beds_filter or u.get('bedrooms') == beds_int)
                      and not u.get('floor')]
            # Es "sin verificar" si no hay strict match (ni con precio igual ni diferente)
            # y no hay loose con precio igual
            if not strict and not any(abs(u['price'] - master_price) < 1 for u in loose):
                n += 1
        return n

    def _count_changes(self, dev_name: str, pdf_units: list[dict]) -> int:
        """Cuenta cuántas filas del Excel para este desarrollo tendrían cambio."""
        if 'Development' not in self.df.columns:
            return 0
        dev_rows = self.df[self.df['Development'] == dev_name]
        is_ocr   = any(u.get('ocr') for u in pdf_units)
        used_ids: set = set()
        n = 0
        for _, row in dev_rows.iterrows():
            master_price = parse_price(row.get('Price'))
            if master_price is None:
                continue
            try:
                beds_raw = row.get('Bedrooms')
                beds_int = int(float(str(beds_raw))) if (
                    beds_raw is not None and str(beds_raw) not in ('', 'nan', '—')
                ) else None
            except (ValueError, TypeError):
                beds_int = None
            floor_cat  = classify_floor(str(row.get('Floor', '')))
            master_ref = str(row.get('Reference', '') or '')

            if is_ocr:
                avail_ocr   = [u for u in pdf_units
                               if u.get('status') == 'AVAILABLE' and u.get('price')]
                beds_filter = (beds_int is None)
                strict = [u for u in avail_ocr
                          if (beds_filter or u.get('bedrooms') == beds_int)
                          and u.get('floor')
                          and (classify_floor(u['floor']) == floor_cat or floor_cat == 'other')]
                loose  = [u for u in avail_ocr
                          if (beds_filter or u.get('bedrooms') == beds_int)
                          and not u.get('floor')]
                if strict and not any(abs(u['price'] - master_price) < 1 for u in strict):
                    if not any(abs(u['price'] - master_price) < 1 for u in loose):
                        n += 1
            else:
                action, _ = self._match_row_to_pdf(
                    master_price, master_ref, beds_int, floor_cat, pdf_units, used_ids)
                if action != 'no_change':
                    n += 1
        return n

    def _apply_auto_update(self, matched_results: list[dict], dlg: tk.Toplevel):
        """Aplica los cambios detectados al DataFrame y refresca la tabla."""
        n_changed  = 0
        n_sold_out = 0
        manual     = []

        for result in matched_results:
            dev_name  = result.get('dev')
            pdf_units = result.get('units', [])

            if not dev_name or not pdf_units:
                if result.get('tag') in ('nomatch', 'manual'):
                    manual.append(result['pdf'].name)
                continue

            is_ocr   = any(u.get('ocr') for u in pdf_units)
            used_ids: set = set()

            dev_mask = (self.df['Development'] == dev_name) if 'Development' in self.df.columns else pd.Series(False, index=self.df.index)

            for idx in self.df[dev_mask].index:
                row = self.df.loc[idx]
                master_price = parse_price(row.get('Price'))
                if master_price is None:
                    continue

                try:
                    beds_raw = row.get('Bedrooms')
                    bedrooms = int(float(str(beds_raw))) if (
                        beds_raw is not None and str(beds_raw) not in ('', 'nan', '—')
                    ) else None
                except (ValueError, TypeError):
                    bedrooms = None
                floor_cat  = classify_floor(str(row.get('Floor', '')))
                master_ref = str(row.get('Reference', '') or '')

                if is_ocr:
                    # ── PDF de imagen (OCR): matching de dos niveles ───────────
                    # Nivel 1 (estricto): mismas camas + planta confirmada
                    # Nivel 2 (suelto): mismas camas + planta desconocida por OCR
                    # Solo aplicamos cambio si nivel 1 confirma precio distinto.
                    avail_ocr   = [u for u in pdf_units
                                   if u.get('status') == 'AVAILABLE' and u.get('price')]
                    beds_filter = (bedrooms is None)
                    strict = [u for u in avail_ocr
                              if (beds_filter or u.get('bedrooms') == bedrooms)
                              and u.get('floor')
                              and (classify_floor(u['floor']) == floor_cat or floor_cat == 'other')]
                    loose  = [u for u in avail_ocr
                              if (beds_filter or u.get('bedrooms') == bedrooms)
                              and not u.get('floor')]

                    if strict:
                        if any(abs(u['price'] - master_price) < 1 for u in strict):
                            continue  # Sin cambio (planta + precio confirmados)
                        if any(abs(u['price'] - master_price) < 1 for u in loose):
                            continue  # Precio coincide en unidad sin planta (OCR no leyó planta)
                        # Precio distinto en strict Y no hay loose con precio correcto → cambio real
                        best      = min(strict, key=lambda u: u['price'])
                        new_price = best['price']
                    elif any(abs(u['price'] - master_price) < 1 for u in loose):
                        continue  # Precio coincide aunque OCR no leyó la planta
                    else:
                        # No encontrado o precio distinto sin planta → manual
                        manual.append(f"{row.get('Reference','?')} ({row.get('Development','?')}) — OCR: planta o precio no confirmado")
                        continue

                    new_price = new_price  # ya asignado arriba (sólo llega aquí si strict y precio distinto)
                    self.df.at[idx, 'Price']   = str(int(new_price)) if new_price == int(new_price) else str(new_price)
                    self.df.at[idx, '_status'] = STATUS_CHANGED
                    self.df.at[idx, '_change'] = (
                        f'Precio (OCR): {master_price:,.0f}€ → {new_price:,.0f}€  '
                        f'(Δ {new_price - master_price:+,.0f}€)'
                    )
                    self.changes.append({
                        'type':        'price_change',
                        'promotor':    row.get('Promotor', ''),
                        'development': dev_name,
                        'reference':   row.get('Reference', ''),
                        'old_price':   master_price,
                        'new_price':   new_price,
                        'file':        result['pdf'].name,
                        'idx':         idx,
                    })
                    n_changed += 1

                else:
                    # ── PDF de texto: matching por referencia (o precio como fallback) ──
                    action, pdf_unit = self._match_row_to_pdf(
                        master_price, master_ref, bedrooms, floor_cat, pdf_units, used_ids)

                    if action == 'no_change':
                        continue

                    elif action in ('price_change', 'replacement'):
                        new_price = pdf_unit['price']
                        kind_lbl  = (
                            '🔄 Sustitución' if action == 'replacement'
                            else '🟡 Precio cambiado'
                        )
                        self.df.at[idx, 'Price']   = str(int(new_price)) if new_price == int(new_price) else str(new_price)
                        self.df.at[idx, '_status'] = STATUS_CHANGED
                        self.df.at[idx, '_change'] = (
                            f'{kind_lbl}: {master_price:,.0f}€ → {new_price:,.0f}€  '
                            f'(Δ {new_price - master_price:+,.0f}€)'
                        )
                        if action == 'replacement':
                            new_ref = pdf_unit.get('reference', '')
                            if new_ref and new_ref != master_ref:
                                self.df.at[idx, 'Reference'] = new_ref
                        self.changes.append({
                            'type':        action,
                            'promotor':    row.get('Promotor', ''),
                            'development': dev_name,
                            'reference':   row.get('Reference', ''),
                            'new_ref':     pdf_unit.get('reference', ''),
                            'old_price':   master_price,
                            'new_price':   new_price,
                            'file':        result['pdf'].name,
                            'idx':         idx,
                        })
                        n_changed += 1

                    else:  # sold_out
                        self.df.at[idx, '_status']      = STATUS_SOLD
                        self.df.at[idx, 'Observations'] = 'Categoría agotada — archivar en Odoo'
                        self.df.at[idx, '_change']      = 'Sin reemplazo disponible en el PDF'
                        self.changes.append({
                            'type':        'sold_out',
                            'promotor':    row.get('Promotor', ''),
                            'development': dev_name,
                            'reference':   row.get('Reference', ''),
                            'price':       master_price,
                            'file':        result['pdf'].name,
                            'idx':         idx,
                        })
                        n_sold_out += 1

        self._refresh_table()
        self._refresh_promotor_list()
        self._update_stats()
        dlg.destroy()

        lines = ['✅  Actualización completada:\n']
        if n_changed:
            lines.append(f'🟡  {n_changed} fila(s) con cambio de precio (amarillo)')
        if n_sold_out:
            lines.append(f'⬜  {n_sold_out} categoría(s) agotada(s) (gris)')
        if not n_changed and not n_sold_out:
            lines.append('✅  Sin cambios — todos los precios coinciden')
        if manual:
            lines.append(f'\n⚠️  {len(manual)} PDF(s) requieren revisión manual:')
            for f in manual:
                lines.append(f'    • {f}')

        messagebox.showinfo('Actualización aplicada', '\n'.join(lines))
        self._status_var.set(
            f'Actualizado — {n_changed} cambios, {n_sold_out} agotadas')

    # ── Export Excel ─────────────────────────────────────────────

    def cmd_export_excel(self):
        if self.df.empty:
            messagebox.showinfo('Sin datos', 'No hay datos para exportar.')
            return
        today = datetime.now().strftime('%Y-%m-%d')
        default_name = f'KH_{today}.xlsx'
        initial_dir = str(Path(self.current_file).parent) if self.current_file else str(SCRIPT_DIR)

        path = filedialog.asksaveasfilename(
            title='Exportar Excel actualizado',
            initialdir=initial_dir,
            initialfile=default_name,
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx')])
        if not path:
            return

        try:
            self._write_excel(path)
            messagebox.showinfo('Exportado', f'Excel guardado:\n{path}')
            self._status_var.set(f'Exportado: {Path(path).name}')
        except Exception as e:
            messagebox.showerror('Error al exportar', str(e))

    def _write_excel(self, path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'KH'

        header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        data_font   = Font(name='Arial', size=10)
        center      = Alignment(horizontal='center', vertical='center')
        left        = Alignment(horizontal='left',   vertical='center', wrap_text=False)

        thin = Border(bottom=Side(style='thin', color='CCCCCC'))

        # Header row
        for ci, col in enumerate(MASTER_COLUMNS, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font      = header_font
            c.fill      = FILL_HEADER
            c.alignment = center
        ws.row_dimensions[1].height = 24

        # Find the column index for 'Price' (for per-cell red fill)
        price_col_idx = MASTER_COLUMNS.index('Price') + 1 if 'Price' in MASTER_COLUMNS else None

        # Sort: sold rows go to bottom in Excel too
        export_df = self.df.copy()
        if '_status' in export_df.columns:
            export_df['_sort_key'] = export_df['_status'].map(
                lambda s: STATUS_ORDER.get(str(s), 0))
            export_df = export_df.sort_values('_sort_key', kind='stable').drop(columns=['_sort_key'])

        # Data rows
        for ri, (_, row) in enumerate(export_df.iterrows(), 2):
            status = str(row.get('_status', ''))

            if status == STATUS_CHANGED:
                row_fill = FILL_YELLOW
            elif status == STATUS_SOLD:
                row_fill = FILL_GRAY
            else:
                row_fill = None  # no_price and normal: no full-row fill

            for ci, col in enumerate(MASTER_COLUMNS, 1):
                v = row.get(col, '')
                if pd.isna(v) or str(v) == 'nan':
                    v = ''
                cell = ws.cell(row=ri, column=ci, value=str(v) if v else '')
                cell.font      = data_font
                cell.alignment = left
                cell.border    = thin

                if row_fill:
                    cell.fill = row_fill
                elif status == STATUS_NO_PRICE and ci == price_col_idx:
                    # Only color the Price cell in red for no_price rows
                    cell.fill = FILL_RED_CELL
                    cell.font = Font(name='Arial', size=10, color='C62828')

        # Column widths
        for ci, col in enumerate(MASTER_COLUMNS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS_XL.get(col, 14)

        ws.freeze_panes      = 'A2'
        ws.auto_filter.ref   = ws.dimensions
        ws.sheet_view.showGridLines = True

        wb.save(path)

    # ── Report ───────────────────────────────────────────────────

    def cmd_generate_report(self):
        if self.df.empty:
            messagebox.showinfo('Sin datos', 'No hay datos para generar informe.')
            return
        today = datetime.now().strftime('%Y-%m-%d')
        initial_dir = str(Path(self.current_file).parent) if self.current_file else str(SCRIPT_DIR)

        path = filedialog.asksaveasfilename(
            title='Guardar informe de cambios',
            initialdir=initial_dir,
            initialfile=f'Informe_Cambios_{today}.md',
            defaultextension='.md',
            filetypes=[('Markdown', '*.md'), ('Texto', '*.txt')])
        if not path:
            return

        try:
            self._write_report(path, today)
            messagebox.showinfo('Informe generado', f'Informe guardado:\n{path}')
        except Exception as e:
            messagebox.showerror('Error al generar informe', str(e))

    def _write_report(self, path: str, date: str):
        s = self.df.get('_status', pd.Series(dtype=str)) if '_status' in self.df.columns else pd.Series(dtype=str)
        changed_df  = self.df[s == STATUS_CHANGED]  if not s.empty else pd.DataFrame()
        no_price_df = self.df[s == STATUS_NO_PRICE] if not s.empty else pd.DataFrame()
        sold_df     = self.df[s == STATUS_SOLD]     if not s.empty else pd.DataFrame()

        lines = [
            f'# Informe de cambios — Listado maestro de propiedades',
            f'**Fecha:** {date}',
            f'**Archivo fuente:** `{Path(self.current_file).name if self.current_file else "—"}`',
            '',
            '---',
            '',
            '## 1. Cambios de precio / sustituciones (filas amarillas)',
            '',
        ]

        if not changed_df.empty:
            lines += [
                '| Promotor | Promoción | Referencia | Dorm | Planta | Precio nuevo | Cambio |',
                '|---|---|---|---|---|---|---|',
            ]
            for _, r in changed_df.iterrows():
                lines.append(
                    f"| {r.get('Promotor','')} "
                    f"| {r.get('Development','')} "
                    f"| {r.get('Reference','')} "
                    f"| {r.get('Bedrooms','')} "
                    f"| {r.get('Floor','')} "
                    f"| {r.get('Price','')} "
                    f"| {r.get('_change','')} |"
                )
        else:
            lines.append('_No se registraron cambios de precio en esta sesión._')

        lines += [
            '',
            '---',
            '',
            '## 2. Sin precio listado (celda precio en rojo)',
            '',
        ]

        if not no_price_df.empty:
            lines += [
                '| Promotor | Promoción | Referencia | Dorm | Planta | Observaciones |',
                '|---|---|---|---|---|---|',
            ]
            for _, r in no_price_df.iterrows():
                lines.append(
                    f"| {r.get('Promotor','')} "
                    f"| {r.get('Development','')} "
                    f"| {r.get('Reference','')} "
                    f"| {r.get('Bedrooms','')} "
                    f"| {r.get('Floor','')} "
                    f"| {r.get('Observations','Pendiente de precio')} |"
                )
        else:
            lines.append('_No hay propiedades sin precio._')

        lines += [
            '',
            '---',
            '',
            '## 3. Vendidas / no disponibles (filas grises — archivar en Odoo)',
            '',
        ]

        if not sold_df.empty:
            lines += [
                '| Promotor | Promoción | Referencia | Dorm | Categoría | Observaciones |',
                '|---|---|---|---|---|---|',
            ]
            for _, r in sold_df.iterrows():
                lines.append(
                    f"| {r.get('Promotor','')} "
                    f"| {r.get('Development','')} "
                    f"| {r.get('Reference','')} "
                    f"| {r.get('Bedrooms','')} "
                    f"| {r.get('Floor','')} "
                    f"| {r.get('Observations','')} |"
                )
        else:
            lines.append('_No hay categorías marcadas como vendidas/agotadas._')

        total    = len(self.df)
        n_chg    = len(changed_df)
        n_nopx   = len(no_price_df)
        n_sold   = len(sold_df)
        n_ok     = total - n_chg - n_nopx - n_sold

        lines += [
            '',
            '---',
            '',
            '## 4. Resumen ejecutivo',
            '',
            f'- **Total propiedades en el maestro:** {total}',
            f'- **Precio modificado (amarillo):** {n_chg}',
            f'- **Sin precio (rojo):** {n_nopx}',
            f'- **Vendidas / agotadas (gris, archivar en Odoo):** {n_sold}',
            f'- **Sin cambios:** {n_ok}',
            '',
            '---',
            f'_Generado automaticamente por Property Manager v{VERSION} — {date}_',
        ]

        with open(path, 'w', encoding='utf-8') as fh:
            fh.write('\n'.join(lines))


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    root = tk.Tk()
    try:
        # Try to set a nice window icon (optional)
        root.iconbitmap(default='')
    except Exception:
        pass
    app = PropertyManagerApp(root)  # noqa: F841
    root.mainloop()


if __name__ == '__main__':
    main()
